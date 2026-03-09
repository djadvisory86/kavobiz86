"""
☕ КавоБот — облік кавового бізнесу Володимир + Коля
Запуск: python bot.py
Потрібно: pip install python-telegram-bot aiohttp openpyxl
"""
import json, os, re, logging, tempfile
from datetime import datetime, timedelta
from telegram import Update, InlineKeyboardButton, InlineKeyboardMarkup, ReplyKeyboardMarkup
from telegram.ext import Application, CommandHandler, MessageHandler, CallbackQueryHandler, ContextTypes, filters
import aiohttp

logging.basicConfig(format="%(asctime)s - %(levelname)s - %(message)s", level=logging.INFO)
logger = logging.getLogger(__name__)

# ══════════════════════════════════════════════════════
# КОНФІГ — ЗАМІНИ ПЕРЕД ЗАПУСКОМ
# ══════════════════════════════════════════════════════
BOT_TOKEN    = os.environ.get("BOT_TOKEN",    "8414849953:AAFeewGPh0BNSWhdY5jGkNdVgFeWVVt51sU")
GROQ_API_KEY = os.environ.get("GROQ_API_KEY", "ВАШ_GROQ_KEY")   # console.groq.com — безкоштовно

ID_VOLODYMYR   = int(os.environ.get("ID_VOLODYMYR",   "373296886"))   # Володимир
ID_VYGRAN      = int(os.environ.get("ID_VYGRAN",      "987654321"))   # Коля — заміни!
ID_GROUP_CHAT  = int(os.environ.get("ID_GROUP_CHAT",  "0"))           # Спільна група (0 = вимкнено)

DEBT_ALERT   = 15000   # нагадування якщо борг більше цієї суми
CONFIRM_SUM  = 5000    # підтвердження обома якщо сума більше

# Нагадування про оренду
RENT_DAY     = int(os.environ.get("RENT_DAY", "1"))    # день місяця (1 = перше)
RENT_AMOUNT  = int(os.environ.get("RENT_AMOUNT", "0")) # сума оренди (0 = вимкнено)

DATA_FILE  = "data.json"
PIN_FILE   = "pins.json"
EXPORT_DIR = "exports"

# ══════════════════════════════════════════════════════
# ЦІНИ
# ══════════════════════════════════════════════════════
PRICES = {
    "кава":      {"buy": 530, "sell": 700, "my": 530, "his": 630},
    "комплект":  {"buy": 680, "sell": 850, "my": 680, "his": 780},
    "молоко":    {"buy": 0,   "sell": 0,   "my": 0,   "his": 100},
    "айріш":     {"buy": 0,   "sell": 0,   "my": 0,   "his": 100},
    "стакан110": {"buy": 0,   "sell": 0,   "my": 0,   "his": 0},
    "стакан250": {"buy": 0,   "sell": 0,   "my": 0,   "his": 0},
}

DEFAULT_POINTS = {}  # Точки додаються через бота командою /points

# ══════════════════════════════════════════════════════
# БАЗА ДАНИХ
# ══════════════════════════════════════════════════════
def load() -> dict:
    if os.path.exists(DATA_FILE):
        with open(DATA_FILE, "r", encoding="utf-8") as f:
            return json.load(f)
    return {
        "points": DEFAULT_POINTS.copy(),
        "transactions": [],
        "stock": {"кава": 0, "комплект": 0, "молоко": 0, "айріш": 0, "стакан110": 0, "стакан250": 0},
        "balance": 0.0,
        # balance > 0 = Коля винен Володимиру
        # balance < 0 = Володимир винен Колі
        "weekly_sent": "",
        "stats": {
            "my":  {"кава": 0, "комплект": 0, "молоко": 0, "айріш": 0},
            "his": {"кава": 0, "комплект": 0, "молоко": 0, "айріш": 0},
        },
        "clients": {},
        "price_history": [],
    }

def save(d: dict):
    with open(DATA_FILE, "w", encoding="utf-8") as f:
        json.dump(d, f, ensure_ascii=False, indent=2)

def load_pins() -> dict:
    if os.path.exists(PIN_FILE):
        with open(PIN_FILE) as f:
            return json.load(f)
    return {}

def save_pins(p: dict):
    with open(PIN_FILE, "w") as f:
        json.dump(p, f)

def add_tx(d: dict, ttype: str, desc: str, amount: float, delta: float, meta: dict = None):
    tx = {
        "id": len(d["transactions"]) + 1,
        "date": datetime.now().strftime("%Y-%m-%d %H:%M"),
        "type": ttype, "description": desc,
        "amount": round(amount, 2),
        "balance_delta": round(delta, 2),
        "meta": meta or {},
    }
    d["transactions"].append(tx)
    d["balance"] = round(d.get("balance", 0) + delta, 2)
    save(d)
    return tx

# ══════════════════════════════════════════════════════
# ДОПОМІЖНІ ФУНКЦІЇ
# ══════════════════════════════════════════════════════
def ok(uid): return uid in (ID_VOLODYMYR, ID_VYGRAN) or (ID_GROUP_CHAT != 0 and uid == ID_GROUP_CHAT)

async def notify_all(bot, msg: str, parse_mode="Markdown", kb=None):
    """Надсилає повідомлення в групу якщо є, або обом особисто"""
    targets = []
    if ID_GROUP_CHAT:
        targets = [ID_GROUP_CHAT]
    else:
        targets = [ID_VOLODYMYR, ID_VYGRAN]
    for t in targets:
        try:
            await bot.send_message(t, msg, parse_mode=parse_mode, reply_markup=kb)
        except Exception as e:
            logger.warning(f"notify_all failed for {t}: {e}")
def uname(uid): return "Володимир" if uid == ID_VOLODYMYR else "Коля"
def other_id(uid): return ID_VYGRAN if uid == ID_VOLODYMYR else ID_VOLODYMYR
def fm(v): return f"{abs(v):,.0f} грн".replace(",", " ")

def get_prices(d: dict) -> dict:
    """Повертає актуальні ціни (з можливістю перевизначення в data.json)"""
    base = {
        "кава":      {"buy": 530, "sell": 700, "my": 530, "his": 630},
        "комплект":  {"buy": 680, "sell": 850, "my": 680, "his": 780},
        "молоко":    {"buy": 0,   "sell": 0,   "my": 0,   "his": 100},
        "айріш":     {"buy": 0,   "sell": 0,   "my": 0,   "his": 100},
        "стакан110": {"buy": 0,   "sell": 0,   "my": 0,   "his": 0},
        "стакан250": {"buy": 0,   "sell": 0,   "my": 0,   "his": 0},
    }
    override = d.get("prices", {})
    for good in base:
        if good in override:
            base[good].update(override[good])
    return base

def bal_line(b: float) -> str:
    if abs(b) < 1: return "✅ Рахунки зведені"
    if b > 0: return f"🔴 Коля винен Володимиру: *{fm(b)}*"
    return f"🔴 Володимир винен Колі: *{fm(b)}*"

def find_point(name: str, points: dict) -> tuple[str, str] | tuple[None, None]:
    name_l = name.lower().strip()
    for p, owner in points.items():
        if name_l == p.lower() or name_l in p.lower() or p.lower() in name_l:
            return p, owner
    return None, None

# ══════════════════════════════════════════════════════
# СТАН ДІАЛОГІВ
# ══════════════════════════════════════════════════════
PIN_WAIT   = {}   # uid -> "set" | "check"
QUICK      = {}   # uid -> {step, ...}
PENDING    = {}   # key -> {desc, amount, confirmed_by, payer, comment}

# ══════════════════════════════════════════════════════
# AI — GROQ (безкоштовно)
# ══════════════════════════════════════════════════════
AI_PROMPT = """Ти асистент для обліку кавового бізнесу. Розпізнай операцію і поверни ТІЛЬКИ JSON.

Точки Володимира: Мурчик, Вдов, Сухопара, Гордівка, Торканівка
Точки Колі: Оляниця, Мамина вишня, Клуб, Ася, Ободівка агро, Ковалівка, Корпуса

Типи: sale=продаж, supply=постачання товару, payment=виплата грошей,
      rent=оренда терміналів, equipment=обладнання, delivery=доставка, unknown=незрозуміло

Товари: кава, комплект, молоко, айріш, стакан110, стакан250

Поверни JSON (без markdown):
{
  "type": "sale|supply|payment|rent|equipment|delivery|unknown",
  "point": "назва або null",
  "items": {"кава": 5},
  "amount": 1500,
  "payer": "volodymyr|vygran|null",
  "comment": "опис"
}"""

async def ai_parse(text: str) -> dict | None:
    if not GROQ_API_KEY or "GROQ" in GROQ_API_KEY:
        return None
    try:
        async with aiohttp.ClientSession() as s:
            async with s.post(
                "https://api.groq.com/openai/v1/chat/completions",
                headers={"Authorization": f"Bearer {GROQ_API_KEY}", "Content-Type": "application/json"},
                json={
                    "model": "llama3-8b-8192",
                    "messages": [{"role": "system", "content": AI_PROMPT}, {"role": "user", "content": text}],
                    "temperature": 0.1, "max_tokens": 200,
                }
            ) as r:
                if r.status == 200:
                    data = await r.json()
                    raw = data["choices"][0]["message"]["content"].strip()
                    raw = re.sub(r"```json|```", "", raw).strip()
                    return json.loads(raw)
    except Exception as e:
        logger.error(f"AI: {e}")
    return None

async def transcribe(path: str) -> str | None:
    if not GROQ_API_KEY or "GROQ" in GROQ_API_KEY:
        return None
    try:
        async with aiohttp.ClientSession() as s:
            with open(path, "rb") as f:
                form = aiohttp.FormData()
                form.add_field("file", f, filename="audio.ogg", content_type="audio/ogg")
                form.add_field("model", "whisper-large-v3")
                form.add_field("language", "uk")
                async with s.post(
                    "https://api.groq.com/openai/v1/audio/transcriptions",
                    headers={"Authorization": f"Bearer {GROQ_API_KEY}"},
                    data=form
                ) as r:
                    if r.status == 200:
                        res = await r.json()
                        return res.get("text", "")
    except Exception as e:
        logger.error(f"Voice: {e}")
    return None

# ══════════════════════════════════════════════════════
# ОБРОБКА ОПЕРАЦІЙ
# ══════════════════════════════════════════════════════
def make_receipt(point: str, items: dict, prices: dict, now: str = None) -> str:
    if not now:
        now = datetime.now().strftime("%d.%m.%Y %H:%M")
    lines = [
        "━━━━━━━━━━━━━━━━━━━━",
        "       ☕ КавоМаркет",
        f"  {now}",
        f"  Точка: {point}",
        "━━━━━━━━━━━━━━━━━━━━",
    ]
    total = 0
    for good, qty in items.items():
        p    = prices.get(good, {})
        sell = p.get("sell", 0)
        sub  = sell * qty
        total += sub
        lines.append(f"  {good:12} {qty}шт × {sell} = {sub} грн")
    lines += [
        "━━━━━━━━━━━━━━━━━━━━",
        f"  РАЗОМ:       {total} грн",
        "━━━━━━━━━━━━━━━━━━━━",
        "  Дякуємо! ☕",
    ]
    return "\n".join(lines)

async def do_sale(point: str, owner: str, items: dict, d: dict) -> str:
    prices = get_prices(d)
    lines  = [f"☕ *Продаж — {point}* ({'моя' if owner == 'volodymyr' else 'Колі'})\n"]
    total_pay = 0
    total_rev = 0
    for good, qty in items.items():
        p = prices.get(good)
        if not p:
            continue
        sell = p.get("sell", 0)
        pay  = p["my"] if owner == "volodymyr" else p["his"]
        rev  = sell * qty
        total_pay += pay * qty
        total_rev += rev
        lines.append(f"  {good}: {qty} шт × {sell} = {fm(rev)} → Колі: {fm(pay*qty)}")
        if good in d["stock"]:
            d["stock"][good] = max(0, d["stock"][good] - qty)
        sk = "my" if owner == "volodymyr" else "his"
        if good in d["stats"][sk]:
            d["stats"][sk][good] += qty

    lines.append(f"\n💰 До виплати Колі: *{fm(total_pay)}*")
    lines.append(f"💵 Виручка: {fm(total_rev)}")
    lines.append(f"📈 Твій заробіток: *{fm(total_rev - total_pay)}*")
    add_tx(d, "sale", f"Продаж {point}: {items}", total_pay, -total_pay,
           {"point": point, "owner": owner, "items": items})
    for good, qty in items.items():
        rem = d["stock"].get(good, 0)
        if rem <= 3:
            lines.append(f"⚠️ {good}: залишилось лише {rem} шт!")
    d["_last_receipt"] = make_receipt(point, items, prices)
    d["_last_sale"]    = {"point": point, "items": items}

    # Перевіряємо борги клієнтів — якщо хтось має борг, нагадуємо
    clients_with_debt = [(n, c["debt"]) for n, c in d.get("clients", {}).items() if c.get("debt", 0) > 0]
    if clients_with_debt:
        lines.append("")
        lines.append("👥 *Клієнти з боргами:*")
        for cname, cdebt in sorted(clients_with_debt, key=lambda x: -x[1])[:3]:
            lines.append(f"  • {cname}: {fm(cdebt)}")
        if len(clients_with_debt) > 3:
            lines.append(f"  _... ще {len(clients_with_debt)-3} клієнт(ів)_")

    save(d)
    return "\n".join(lines)


async def do_supply(items: dict, amount: float, d: dict, comment: str = "") -> str:
    lines = ["📥 *Постачання від Колі*\n"]
    total = 0
    for good, qty in items.items():
        if good in d["stock"]:
            d["stock"][good] += qty
        buy = PRICES.get(good, {}).get("buy", 0)
        sub = buy * qty
        total += sub
        lines.append(f"  {good}: +{qty} шт" + (f" = {fm(sub)}" if sub else ""))
        lines.append(f"  → Склад тепер: {d['stock'].get(good, qty)} шт")
    if amount:
        total = amount
    if comment:
        lines.append(f"\n💬 {comment}")
    if total:
        lines.append(f"\n💰 Сума постачання: *{fm(total)}*")
    # Отримали товар → більше винні → баланс -
    add_tx(d, "supply", f"Постачання: {items}", total, -total, {"items": items})
    return "\n".join(lines)

async def do_payment(amount: float, payer: str, d: dict, comment: str = "") -> str:
    if payer == "volodymyr":
        desc  = f"💸 Володимир → Колі: *{fm(amount)}*"
        delta = amount   # заплатили → борг зменшився → баланс +
    else:
        desc  = f"💸 Коля → Володимиру: *{fm(amount)}*"
        delta = -amount
    if comment:
        desc += f"\n💬 {comment}"
    add_tx(d, "payment", desc, amount, delta, {"payer": payer})
    return desc

async def do_expense(etype: str, amount: float, payer: str, d: dict, comment: str = "") -> str:
    half = round(amount / 2, 2)
    names = {"rent": "🏠 Оренда", "equipment": "🔧 Обладнання", "delivery": "🚚 Доставка"}
    name  = names.get(etype, "💰 Витрата")
    if payer == "vygran":
        desc  = f"{name}: *{fm(amount)}*\nКоля заплатив → ти йому винен половину ({fm(half)})"
        delta = -half
    else:
        desc  = f"{name}: *{fm(amount)}*\nТи заплатив → Коля тобі винен половину ({fm(half)})"
        delta = half
    if comment:
        desc += f"\n💬 {comment}"
    add_tx(d, etype, desc, amount, delta, {"payer": payer, "amount": amount})
    return desc

# ══════════════════════════════════════════════════════
# КЛАВІАТУРА
# ══════════════════════════════════════════════════════
def main_kb():
    return ReplyKeyboardMarkup([
        ["☕ Продаж кави",        "📦 Продаж комплекту"],
        ["🥛 Молоко / Айріш",    "📥 Постачання"],
        ["💸 Виплата",            "🏠 Оренда / Витрати"],
        ["💰 Баланс",             "📊 Звіт"],
        ["📍 Точки",              "👥 Борги клієнтів"],
        ["💲 Змінити ціни",       "🧾 Чек клієнту"],
        ["📊 Графік",             "🔔 Нагадати клієнтам"],
        ["🤝 Розрахунок з Колею", "↩️ Скасувати"],
        ["⚙️ Налаштування"],
    ], resize_keyboard=True)

# ══════════════════════════════════════════════════════
# КОМАНДИ
# ══════════════════════════════════════════════════════
async def cmd_start(update: Update, ctx: ContextTypes.DEFAULT_TYPE):
    uid = update.effective_user.id
    if not ok(uid):
        await update.message.reply_text("⛔ Доступ заборонено.")
        return
    await update.message.reply_text(
        f"☕ Привіт, {uname(uid)}!\n\n"
        "Пиши як у групі: _«Оляниця 5 кав»_\n"
        "Або надішли голосове повідомлення\n"
        "Або використовуй кнопки нижче 👇",
        reply_markup=main_kb(), parse_mode="Markdown"
    )

async def cmd_balance(update: Update, ctx: ContextTypes.DEFAULT_TYPE):
    uid = update.effective_user.id
    if not ok(uid): return
    pins = load_pins()
    if str(uid) in pins:
        PIN_WAIT[uid] = "check"
        await update.message.reply_text("🔐 Введи PIN:")
    else:
        await show_balance(update)

async def show_balance(update: Update):
    d = load()
    b = d.get("balance", 0)
    s = d.get("stock", {})
    sm = d["stats"]["my"]
    sh = d["stats"]["his"]
    text = (
        f"💰 *БАЛАНС*\n{'─'*26}\n{bal_line(b)}\n\n"
        f"📦 *СКЛАД*\n"
        f"  ☕ Кава: {s.get('кава',0)} шт\n"
        f"  📦 Комплекти: {s.get('комплект',0)} шт\n"
        f"  🥛 Молоко: {s.get('молоко',0)} шт\n"
        f"  🍹 Айріш: {s.get('айріш',0)} шт\n"
        f"  🥤 Ст.110: {s.get('стакан110',0)} шт  "
        f"🥤 Ст.250: {s.get('стакан250',0)} шт\n\n"
        f"📊 *МОЇ ТОЧКИ продано:* кава {sm['кава']} | компл {sm['комплект']}\n"
        f"📊 *ТОЧКИ ВИГРНА продано:* кава {sh['кава']} | компл {sh['комплект']} | "
        f"молоко/айріш {sh['молоко']+sh['айріш']}"
    )
    await update.message.reply_text(text, parse_mode="Markdown")

async def cmd_report(update: Update, ctx: ContextTypes.DEFAULT_TYPE):
    uid = update.effective_user.id
    if not ok(uid): return
    d    = load()
    now  = datetime.now()
    mo   = now.strftime("%Y-%m")
    txs  = [t for t in d["transactions"] if t["date"][:7] == mo]
    sm   = d["stats"]["my"]
    sh   = d["stats"]["his"]

    profit = (sm["кава"]*170 + sm["комплект"]*170 +
              sh["кава"]*70  + sh["комплект"]*70  +
              sh["молоко"]*100 + sh["айріш"]*100)

    sales_sum = sum(t["amount"] for t in txs if t["type"] == "sale")
    sup_sum   = sum(t["amount"] for t in txs if t["type"] == "supply")
    pay_sum   = sum(t["amount"] for t in txs if t["type"] == "payment")
    exp_sum   = sum(t["amount"] for t in txs if t["type"] in ("rent","equipment","delivery"))

    text = (
        f"📊 *ЗВІТ {now.strftime('%B %Y').upper()}*\n{'═'*28}\n\n"
        f"☕ *МОЇ ТОЧКИ:*\n"
        f"  Кава: {sm['кава']} шт × 170 = {fm(sm['кава']*170)}\n"
        f"  Комплекти: {sm['комплект']} шт × 170 = {fm(sm['комплект']*170)}\n\n"
        f"☕ *ТОЧКИ ВИГРНА:*\n"
        f"  Кава: {sh['кава']} шт × 70 = {fm(sh['кава']*70)}\n"
        f"  Комплекти: {sh['комплект']} шт × 70 = {fm(sh['комплект']*70)}\n"
        f"  Молоко+Айріш: {sh['молоко']+sh['айріш']} шт × 100 = "
        f"{fm((sh['молоко']+sh['айріш'])*100)}\n\n"
        f"💚 *Твій прибуток: {fm(profit)}*\n"
        f"{'─'*28}\n"
        f"📥 Постачань: {fm(sup_sum)}\n"
        f"💸 Виплачено: {fm(pay_sum)}\n"
        f"🏠 Витрати: {fm(exp_sum)}\n"
        f"{'═'*28}\n"
        f"{bal_line(d['balance'])}\n\n"
        f"📦 Склад: кава {d['stock'].get('кава',0)} | "
        f"компл {d['stock'].get('комплект',0)} | "
        f"молоко {d['stock'].get('молоко',0)} | "
        f"айріш {d['stock'].get('айріш',0)}"
    )
    await update.message.reply_text(text, parse_mode="Markdown")

async def cmd_undo(update: Update, ctx: ContextTypes.DEFAULT_TYPE):
    uid = update.effective_user.id
    if not ok(uid): return
    d   = load()
    txs = d.get("transactions", [])
    if not txs:
        await update.message.reply_text("❌ Немає операцій для скасування.")
        return
    last = txs.pop()
    d["balance"] = round(d["balance"] - last["balance_delta"], 2)
    # Відкат складу
    if last["type"] == "sale":
        for g, q in last.get("meta", {}).get("items", {}).items():
            if g in d["stock"]: d["stock"][g] += q
        sk = "my" if last["meta"].get("owner") == "volodymyr" else "his"
        for g, q in last["meta"].get("items", {}).items():
            if g in d["stats"][sk]:
                d["stats"][sk][g] = max(0, d["stats"][sk][g] - q)
    elif last["type"] == "supply":
        for g, q in last.get("meta", {}).get("items", {}).items():
            if g in d["stock"]: d["stock"][g] = max(0, d["stock"][g] - q)
    save(d)
    await update.message.reply_text(
        f"↩️ Скасовано: _{last['description'][:80]}_\n"
        f"Дата: {last['date']}\n{bal_line(d['balance'])}",
        parse_mode="Markdown"
    )

async def cmd_export(update: Update, ctx: ContextTypes.DEFAULT_TYPE):
    uid = update.effective_user.id
    if not ok(uid): return
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill
        from openpyxl.chart import LineChart, Reference
    except ImportError:
        await update.message.reply_text("❌ pip install openpyxl")
        return
    d   = load()
    txs = d.get("transactions", [])
    wb  = openpyxl.Workbook()
    ws  = wb.active
    ws.title = "Транзакції"
    hdrs = ["ID","Дата","Тип","Опис","Сума","Δ Баланс","Баланс"]
    for c, h in enumerate(hdrs, 1):
        cell = ws.cell(1, c, h)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="6B4226")
    rb = 0
    for row, t in enumerate(txs, 2):
        rb += t.get("balance_delta", 0)
        for c, v in enumerate([t.get("id"), t.get("date"), t.get("type"),
                                t.get("description","")[:50], t.get("amount"),
                                t.get("balance_delta"), round(rb,2)], 1):
            ws.cell(row, c, v)
    for col in ws.columns:
        ws.column_dimensions[col[0].column_letter].width = min(
            max(len(str(c.value or "")) for c in col) + 2, 40)
    # Лист балансу + графік
    ws2 = wb.create_sheet("Баланс")
    ws2.append(["Дата","Баланс"])
    rb = 0
    for t in txs:
        rb += t.get("balance_delta", 0)
        ws2.append([t["date"][:10], round(rb, 2)])
    if len(txs) > 1:
        ch = LineChart()
        ch.title = "Динаміка балансу"
        ch.style = 10
        dr = Reference(ws2, min_col=2, min_row=1, max_row=len(txs)+1)
        ch.add_data(dr, titles_from_data=True)
        ws2.add_chart(ch, "D2")
    os.makedirs(EXPORT_DIR, exist_ok=True)
    fname = f"{EXPORT_DIR}/кавобот_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx"
    wb.save(fname)
    with open(fname, "rb") as f:
        await update.message.reply_document(f, filename=os.path.basename(fname),
            caption="📊 Готово! Відкрий у Excel або Google Sheets.")

async def cmd_setgroup(update: Update, ctx: ContextTypes.DEFAULT_TYPE):
    """Зберігає ID поточного чату як групового"""
    uid = update.effective_user.id
    if uid != ID_VOLODYMYR: return
    chat_id = update.effective_chat.id
    d = load()
    d["group_chat_id"] = chat_id
    save(d)
    # Оновлюємо глобальну змінну
    global ID_GROUP_CHAT
    ID_GROUP_CHAT = chat_id
    await update.message.reply_text(
        f"✅ Цей чат додано як груповий!\n"
        f"ID: `{chat_id}`\n\n"
        f"Тепер всі сповіщення будуть надходити сюди.",
        parse_mode="Markdown")

async def cmd_setpin(update: Update, ctx: ContextTypes.DEFAULT_TYPE):
    uid = update.effective_user.id
    if not ok(uid): return
    PIN_WAIT[uid] = "set"
    await update.message.reply_text("🔐 Введи новий 4-значний PIN:")

async def cmd_points(update: Update, ctx: ContextTypes.DEFAULT_TYPE):
    uid = update.effective_user.id
    if not ok(uid): return
    await show_points_menu(update, ctx)

async def show_points_menu(update, ctx):
    d   = load()
    pts = d.get("points", {})
    if not pts:
        kb = InlineKeyboardMarkup([[
            InlineKeyboardButton("➕ Додати мою точку",    callback_data="padd_vol"),
            InlineKeyboardButton("➕ Додати точку Коли",   callback_data="padd_vyg"),
        ]])
        await update.message.reply_text(
            "📍 *Точок поки немає*\nДодай першу точку:",
            parse_mode="Markdown", reply_markup=kb)
        return

    my  = [(p,o) for p,o in pts.items() if o=="volodymyr"]
    his = [(p,o) for p,o in pts.items() if o=="vygran"]

    lines = ["📍 *УПРАВЛІННЯ ТОЧКАМИ*\n"]
    lines.append("*Мої точки:*")
    for p,_ in my:
        lines.append(f"  • {p}")
    lines.append("\n*Точки Коли:*")
    for p,_ in his:
        lines.append(f"  • {p}")

    kb_rows = []
    # Кнопки для кожної точки
    for p, o in pts.items():
        short = p[:12]+"…" if len(p)>12 else p
        kb_rows.append([
            InlineKeyboardButton(f"✏️ {short}",  callback_data=f"pedit_{p}"),
            InlineKeyboardButton(f"🔄",          callback_data=f"pswap_{p}"),
            InlineKeyboardButton(f"🗑️",          callback_data=f"pdel_{p}"),
        ])
    kb_rows.append([
        InlineKeyboardButton("➕ Додати мою",        callback_data="padd_vol"),
        InlineKeyboardButton("➕ Додати Коли",        callback_data="padd_vyg"),
    ])
    await update.message.reply_text(
        "\n".join(lines),
        parse_mode="Markdown",
        reply_markup=InlineKeyboardMarkup(kb_rows)
    )

async def cmd_addpoint(update: Update, ctx: ContextTypes.DEFAULT_TYPE):
    uid = update.effective_user.id
    if not ok(uid): return
    args = ctx.args
    if len(args) < 2:
        # Інтерактивний режим
        QUICK[uid] = {"step":"newpoint_name", "owner":"volodymyr"}
        await update.message.reply_text(
            "📍 Напиши назву нової точки\n(або /addpoint Назва мій|коля):")
        return
    owner_raw = args[-1].lower()
    owner = "volodymyr" if owner_raw in ("мій","моя","я","vol","volodymyr") else "vygran"
    name  = " ".join(args[:-1])
    d = load()
    d["points"][name] = owner
    save(d)
    who = "моя" if owner == "volodymyr" else "Коли"
    await update.message.reply_text(
        f"✅ Точка *{name}* додана ({who})", parse_mode="Markdown")

async def cmd_stock(update: Update, ctx: ContextTypes.DEFAULT_TYPE):
    uid = update.effective_user.id
    if not ok(uid): return
    d = load()
    s = d.get("stock", {})
    em = {"кава":"☕","комплект":"📦","молоко":"🥛","айріш":"🍹","стакан110":"🥤","стакан250":"🥤"}
    lines = ["📦 *ЗАЛИШКИ НА СКЛАДІ*\n"]
    for g, q in s.items():
        w = " ⚠️ МАЛО!" if q <= 3 else ""
        lines.append(f"  {em.get(g,'•')} {g}: *{q} шт*{w}")
    await update.message.reply_text("\n".join(lines), parse_mode="Markdown")

async def cmd_history(update: Update, ctx: ContextTypes.DEFAULT_TYPE):
    uid = update.effective_user.id
    if not ok(uid): return
    d   = load()
    txs = d.get("transactions",[])[-10:]
    if not txs:
        await update.message.reply_text("📋 Немає операцій.")
        return
    em = {"sale":"☕","supply":"📥","payment":"💸","rent":"🏠","equipment":"🔧","delivery":"🚚"}
    lines = ["📋 *Останні операції:*\n"]
    for t in reversed(txs):
        e = em.get(t["type"],"•")
        sign = "+" if t["balance_delta"] >= 0 else ""
        lines.append(f"{e} *#{t['id']}* {t['date']}\n"
                     f"   {t['description'][:60]}\n"
                     f"   Δ {sign}{fm(t['balance_delta'])}\n")
    await update.message.reply_text("\n".join(lines), parse_mode="Markdown")

async def cmd_settings(update: Update, ctx: ContextTypes.DEFAULT_TYPE):
    uid = update.effective_user.id
    if not ok(uid): return
    await update.message.reply_text(
        "⚙️ *КОМАНДИ*\n\n"
        "/balance — баланс і склад\n"
        "/report — місячний звіт\n"
        "/export — Excel файл\n"
        "/history — останні операції\n"
        "/rozrahunok — розрахунок з Колею\n"
        "/stock — залишки товару\n"
        "/points — список точок\n"
        "/points — управління точками\n"
        "/setpin — PIN для балансу\n"
        "/undo — скасувати останнє\n",
        parse_mode="Markdown"
    )

# ══════════════════════════════════════════════════════
# ПІДТВЕРДЖЕННЯ ВЕЛИКИХ СУМ
# ══════════════════════════════════════════════════════
async def ask_confirm(update: Update, ctx: ContextTypes.DEFAULT_TYPE,
                      key: str, desc: str, amount: float, meta: dict):
    uid   = update.effective_user.id
    other = other_id(uid)
    PENDING[key] = {"desc": desc, "amount": amount, "confirmed_by": {uid}, "meta": meta}
    kb  = InlineKeyboardMarkup([[
        InlineKeyboardButton("✅ Підтвердити", callback_data=f"ok_{key}"),
        InlineKeyboardButton("❌ Відхилити",   callback_data=f"no_{key}"),
    ]])
    msg = (f"⚠️ *Велика сума — потрібно підтвердження обох*\n\n"
           f"{desc}\nСума: *{fm(amount)}*\n\nЧекаю на {uname(other)}...")
    await update.message.reply_text(msg, parse_mode="Markdown", reply_markup=kb)
    try:
        await ctx.bot.send_message(other, msg + f"\n\n_{uname(uid)} ініціював_",
                                   parse_mode="Markdown", reply_markup=kb)
    except Exception:
        pass

# ══════════════════════════════════════════════════════
# CALLBACK КНОПКИ
# ══════════════════════════════════════════════════════
async def on_callback(update: Update, ctx: ContextTypes.DEFAULT_TYPE):
    q   = update.callback_query
    uid = q.from_user.id
    cb  = q.data
    await q.answer()
    if not ok(uid): return
    d = load()

    # Підтвердження/відхилення
    if cb.startswith("ok_") or cb.startswith("no_"):
        key = cb[3:]
        if key not in PENDING:
            await q.edit_message_text("❌ Операція не знайдена (вже оброблена?).")
            return
        if cb.startswith("no_"):
            PENDING.pop(key)
            await q.edit_message_text(f"❌ {uname(uid)} відхилив операцію.")
            try: await ctx.bot.send_message(other_id(uid), f"❌ {uname(uid)} відхилив операцію.")
            except: pass
            return
        PENDING[key]["confirmed_by"].add(uid)
        if len(PENDING[key]["confirmed_by"]) >= 2:
            p = PENDING.pop(key)
            m = p["meta"]
            result = ""
            if m.get("op") == "payment":
                result = await do_payment(p["amount"], m["payer"], d, m.get("comment",""))
            save(d)
            await q.edit_message_text(f"✅ Підтверджено обома!\n\n{result}", parse_mode="Markdown")
            await check_debt(ctx, d)
        else:
            await q.edit_message_text(f"✅ {uname(uid)} підтвердив. Чекаю на {uname(other_id(uid))}...")
        return

    # Вибір точки (швидкий продаж)
    if cb.startswith("pt_"):
        point = cb[3:]
        state = QUICK.get(uid, {})
        good  = state.get("good","кава")
        pending_qty = state.get("pending_qty")
        _, owner = find_point(point, d["points"])
        if "/" in good:
            QUICK[uid] = {"step":"milk_type","point":point,"owner":owner}
            kb = InlineKeyboardMarkup([[
                InlineKeyboardButton("🥛 Молоко", callback_data="mg_молоко"),
                InlineKeyboardButton("🍹 Айріш",  callback_data="mg_айріш"),
            ]])
            await q.edit_message_text(f"📍 {point} — що саме?", reply_markup=kb)
        elif pending_qty:
            QUICK.pop(uid, None)
            result = await do_sale(point, owner, {good: pending_qty}, d)
            await q.edit_message_text(result, parse_mode="Markdown")
            await check_debt(ctx, d)
        else:
            QUICK[uid] = {"step":"qty","point":point,"owner":owner,"good":good}
            await q.edit_message_text(f"📍 {point} — скільки *{good}*? Введи число:", parse_mode="Markdown")
        return

    if cb.startswith("pt2_"):
        parts = cb[4:].rsplit("_", 1)
        if len(parts) == 2:
            point, good = parts
            _, owner = find_point(point, d["points"])
            QUICK[uid] = {"step":"qty","point":point,"owner":owner,"good":good}
            await q.edit_message_text(
                f"📍 *{point}* — скільки *{good}*? Введи число:",
                parse_mode="Markdown")
        return

    # ── ФОТО ЧЕК ──
    if cb.startswith("photo_exp_"):
        amount = float(cb[10:])
        QUICK[uid] = {"step":"exp_payer","etype":"rent","amount":amount}
        kb = InlineKeyboardMarkup([[
            InlineKeyboardButton("Я платив",    callback_data="ep_volodymyr"),
            InlineKeyboardButton("Коля платив", callback_data="ep_vygran"),
        ]])
        await q.edit_message_text(f"💸 Витрата {fm(amount)} — хто платив?", reply_markup=kb)
        return
    if cb.startswith("photo_sup_"):
        amount = float(cb[10:])
        d2 = load()
        result = await do_supply({}, amount, d2, "фото чеку")
        await q.edit_message_text(result, parse_mode="Markdown")
        return
    if cb == "photo_cancel":
        await q.delete_message()
        return

    # ── ЦІНИ ──
    if cb.startswith("sp_"):
        good = cb[3:]
        QUICK[uid] = {"step": "setprice_good", "good": good}
        d2 = load()
        p  = get_prices(d2).get(good, {})
        label = {"кава":"☕ Кава","комплект":"📦 Комплект","молоко":"🥛 Молоко/Айріш"}.get(good, good)
        if good == "молоко":
            await q.edit_message_text(
                f"{label}\nКомісія за продаж на точці Колі зараз: *{p.get('his',100)} грн*\n\nВведи нову комісію:",
                parse_mode="Markdown")
        else:
            await q.edit_message_text(
                f"{label}\nЗакупка: *{p.get('buy')} грн* | Продаж: *{p.get('sell')} грн*\n\n"
                f"Введи через пробіл: _закупка продаж_\nНаприклад: _580 750_",
                parse_mode="Markdown")
        return

    # ── ЧЕК ──
    if cb == "receipt_debt":
        d2   = load()
        sale = d2.get("_last_sale", {})
        QUICK[uid] = {"step": "debt_client_name", "sale": sale}
        await q.edit_message_text("👤 Введи ім'я клієнта (або обери з існуючих):\n\n" +
            "\n".join(f"• {n}" for n in d2.get("clients",{}).keys()) or "Поки немає клієнтів")
        return

    if cb == "show_receipt":
        d2 = load()
        receipt = d2.get("_last_receipt","")
        if receipt:
            await q.edit_message_text(
                f"```\n{receipt}\n```\n_Скопіюй і надішли клієнту_",
                parse_mode="Markdown",
                reply_markup=InlineKeyboardMarkup([[
                    InlineKeyboardButton("📋 У борг клієнту", callback_data="receipt_debt"),
                    InlineKeyboardButton("✅ Закрити",         callback_data="receipt_close"),
                ]]))
        return

    if cb == "receipt_close":
        await q.delete_message()
        return

    # ── КЛІЄНТИ ──
    if cb == "cl_add":
        QUICK[uid] = {"step": "newclient_name"}
        await q.edit_message_text("👤 Введи ім'я нового клієнта:")
        return

    if cb.startswith("cl_view_"):
        name = cb[8:]
        d2   = load()
        c    = d2.get("clients", {}).get(name, {})
        debt = c.get("debt", 0)
        hist = c.get("history", [])[-5:]
        lines = [f"👤 *{name}*", f"Борг: *{fm(debt)}*", ""]
        if hist:
            lines.append("📋 *Останні операції:*")
            for h in reversed(hist):
                sign = "+" if h["type"] == "debt" else "-"
                lines.append(f"  {h['date'][:10]}  {sign}{fm(h['amount'])}  {h.get('note','')}")
        kb = InlineKeyboardMarkup([
            [InlineKeyboardButton("💸 Погасити борг",  callback_data=f"cl_pay_{name}"),
             InlineKeyboardButton("➕ Борг у борг",     callback_data=f"cl_debt_{name}")],
            [InlineKeyboardButton("📋 Чек у борг",     callback_data=f"cl_receipt_{name}"),
             InlineKeyboardButton("🗑️ Видалити",       callback_data=f"cl_del_{name}")],
            [InlineKeyboardButton("◀️ Назад",          callback_data="cl_back")],
        ])
        await q.edit_message_text("\n".join(lines), parse_mode="Markdown", reply_markup=kb)
        return

    if cb == "cl_back":
        d2 = load()
        clients = d2.get("clients", {})
        total   = sum(c.get("debt",0) for c in clients.values())
        lines   = [f"👥 *БОРГИ КЛІЄНТІВ*\nЗагальний борг: *{fm(total)}*\n"]
        kb_rows = []
        for name, c in clients.items():
            short = name[:14]+"…" if len(name)>14 else name
            debt  = c.get("debt",0)
            label = f"{short} ({fm(debt)})" if debt > 0 else short
            kb_rows.append([InlineKeyboardButton(f"👤 {label}", callback_data=f"cl_view_{name}")])
        kb_rows.append([InlineKeyboardButton("➕ Новий клієнт", callback_data="cl_add")])
        await q.edit_message_text("\n".join(lines), parse_mode="Markdown",
                                  reply_markup=InlineKeyboardMarkup(kb_rows))
        return

    if cb.startswith("cl_pay_"):
        name = cb[7:]
        QUICK[uid] = {"step": "client_pay", "client": name}
        d2   = load()
        debt = d2.get("clients",{}).get(name,{}).get("debt",0)
        await q.edit_message_text(
            f"💸 *{name}* погашає борг\nЗараз винен: *{fm(debt)}*\n\nВведи суму оплати:")
        return

    if cb.startswith("cl_debt_"):
        name = cb[8:]
        QUICK[uid] = {"step": "client_add_debt", "client": name}
        await q.edit_message_text(f"➕ *{name}* бере в борг\nВведи суму + опис (необов'язково):\n_Наприклад: 350 2 кави_",
                                  parse_mode="Markdown")
        return

    if cb.startswith("cl_receipt_"):
        name = cb[11:]
        d2   = load()
        sale = d2.get("_last_sale",{})
        prices = get_prices(d2)
        receipt = d2.get("_last_receipt","")
        if not receipt:
            await q.edit_message_text("❌ Немає останнього продажу")
            return
        # Додаємо борг
        items = sale.get("items", {})
        total = sum(prices.get(g,{}).get("sell",0)*q2 for g,q2 in items.items())
        now   = datetime.now()
        c     = d2.setdefault("clients", {}).setdefault(name, {"debt":0,"history":[]})
        c["debt"] = c.get("debt",0) + total
        c["last_date"] = now.isoformat()
        c["history"].append({"type":"debt","amount":total,"date":now.strftime("%d.%m.%Y"),"note":"чек у борг"})
        save(d2)
        await q.edit_message_text(
            f"✅ Борг *{name}* +{fm(total)}\nЗагальний борг: {fm(c['debt'])}",
            parse_mode="Markdown")
        return

    if cb.startswith("cl_del_"):
        name = cb[7:]
        d2 = load()
        d2.get("clients",{}).pop(name, None)
        save(d2)
        await q.edit_message_text(f"🗑️ Клієнта *{name}* видалено.", parse_mode="Markdown")
        return

    # ── УПРАВЛІННЯ ТОЧКАМИ ──
    if cb.startswith("padd_"):
        owner = "volodymyr" if cb == "padd_vol" else "vygran"
        QUICK[uid] = {"step": "newpoint_name", "owner": owner}
        who = "мою" if owner == "volodymyr" else "Коли"
        await q.edit_message_text(f"✏️ Введи назву нової точки ({who}):")
        return

    if cb.startswith("pedit_"):
        point = cb[6:]
        QUICK[uid] = {"step": "rename_point", "old_name": point}
        owner = d["points"].get(point, "volodymyr")
        who = "моя" if owner == "volodymyr" else "Коли"
        await q.edit_message_text(
            f"✏️ *{point}* ({who})\n\nВведи нову назву\n(або надішли крапку . щоб скасувати):",
            parse_mode="Markdown")
        return

    if cb.startswith("pswap_"):
        point = cb[6:]
        old_owner = d["points"].get(point)
        new_owner = "vygran" if old_owner == "volodymyr" else "volodymyr"
        d["points"][point] = new_owner
        save(d)
        who_old = "моя" if old_owner == "volodymyr" else "Коли"
        who_new = "моя" if new_owner == "volodymyr" else "Коли"
        await q.edit_message_text(
            f"🔄 *{point}*\n{who_old} → *{who_new}*\n\n✅ Збережено!",
            parse_mode="Markdown")
        return

    if cb.startswith("pdel_"):
        point = cb[5:]
        kb = InlineKeyboardMarkup([[
            InlineKeyboardButton("✅ Так, видалити", callback_data=f"pdelok_{point}"),
            InlineKeyboardButton("❌ Скасувати",     callback_data="pdelno"),
        ]])
        await q.edit_message_text(
            f"🗑️ Видалити точку *{point}*?\n\n⚠️ Операції з цією точкою залишаться в історії.",
            parse_mode="Markdown", reply_markup=kb)
        return

    if cb.startswith("pdelok_"):
        point = cb[7:]
        if point in d["points"]:
            del d["points"][point]
            save(d)
            await q.edit_message_text(f"🗑️ Точку *{point}* видалено.", parse_mode="Markdown")
        return

    if cb == "pdelno":
        await q.edit_message_text("❌ Скасовано.")
        return

    # ── РОЗРАХУНОК ──
    if cb == "settle_send":
        d2   = load()
        text = build_settlement_text(d2)
        try:
            await ctx.bot.send_message(
                ID_VYGRAN,
                text + "\n\n_Надіслано Колі для підтвердження_",
                parse_mode="Markdown"
            )
            await q.edit_message_text(
                text + "\n\n✅ *Надіслано Колі!*",
                parse_mode="Markdown"
            )
        except Exception as e:
            await q.edit_message_text(f"❌ Не вдалось надіслати: {e}")
        return

    if cb == "settle_paid":
        d2 = load()
        old_bal = d2.get("balance", 0)
        # Записуємо операцію обнулення
        if abs(old_bal) >= 1:
            if old_bal > 0:
                desc = f"🤝 Розрахунок: Коля сплатив {fm(old_bal)}"
                delta = -old_bal
            else:
                desc = f"🤝 Розрахунок: Володимир сплатив {fm(abs(old_bal))}"
                delta = abs(old_bal)
            add_tx(d2, "payment", desc, abs(old_bal), delta, {"settlement": True})
        # Скидаємо статистику продажів (але залишки зберігаємо!)
        d2["stats"] = {
            "my":  {"кава": 0, "комплект": 0, "молоко": 0, "айріш": 0},
            "his": {"кава": 0, "комплект": 0, "молоко": 0, "айріш": 0},
        }
        save(d2)
        # Повідомляємо обох
        msg_done = (
            f"✅ *Розрахунок завершено!*\n\n"
            f"Баланс обнулено.\n"
            f"Залишки на складі збережено.\n"
            f"Статистика продажів скинута.\n\n"
            f"📦 Склад:\n" +
            "\n".join(
                f"  {'☕📦🥛🍹🥤🥤'.split()[i]} {g}: {q2} шт"
                for i,(g,q2) in enumerate(d2["stock"].items()) if q2 > 0
            ) or "  Порожній"
        )
        await q.edit_message_text(msg_done, parse_mode="Markdown")
        try:
            await ctx.bot.send_message(ID_VYGRAN,
                "✅ *Розрахунок підтверджено Володимиром*\n"
                "Баланс обнулено. Наступний період почався.",
                parse_mode="Markdown")
        except Exception:
            pass
        return

    if cb == "settle_close":
        await q.edit_message_text("❌ Розрахунок скасовано.")
        return

    if cb.startswith("mg_"):
        good  = cb[3:]
        state = QUICK.get(uid,{})
        QUICK[uid] = {"step":"qty","point":state["point"],"owner":state["owner"],"good":good}
        await q.edit_message_text(f"📍 {state['point']} — скільки *{good}*? Введи число:", parse_mode="Markdown")
        return

    if cb in ("pay_vol","pay_vyg"):
        payer = "volodymyr" if cb == "pay_vol" else "vygran"
        QUICK[uid] = {"step":"pay_amount","payer":payer}
        who = "Ти → Колі" if payer == "volodymyr" else "Коля → Тобі"
        await q.edit_message_text(f"💸 {who}\nВведи суму (грн):")
        return

    if cb.startswith("ex_"):
        etype = cb[3:]
        QUICK[uid] = {"step":"exp_payer","etype":etype}
        kb = InlineKeyboardMarkup([[
            InlineKeyboardButton("Я платив",      callback_data="ep_volodymyr"),
            InlineKeyboardButton("Коля платив", callback_data="ep_vygran"),
        ]])
        await q.edit_message_text("Хто заплатив?", reply_markup=kb)
        return

    if cb.startswith("ep_"):
        payer = cb[3:]
        state = QUICK.get(uid,{})
        QUICK[uid] = {"step":"exp_amount","etype":state.get("etype","rent"),"payer":payer}
        await q.edit_message_text("Введи суму витрати (грн):")
        return

# ══════════════════════════════════════════════════════
# РОЗУМНИЙ ЛОКАЛЬНИЙ ПАРСИНГ (без AI)
# ══════════════════════════════════════════════════════

# Словник синонімів товарів — всі можливі написання
GOOD_ALIASES = {
    "кава":      ["кав", "кава", "кави", "каву", "кавою", "coffee", "кофе", "коф"],
    "комплект":  ["компл", "комплект", "комплекти", "комплектів", "комплекту",
                  "коплект", "коплектів", "комплектов", "компл.", "комп",
                  "комлпект", "комлпекти", "комлп", "кмплект",
                  "комплекта", "комплектах", "комплектами"],
    "молоко":    ["молок", "молоко", "молока", "молоку"],
    "айріш":     ["айріш", "айриш", "irish", "айрiш"],
    "стакан110": ["ст110", "стакан110", "110мл", "стакани110"],
    "стакан250": ["ст250", "стакан250", "250мл", "стакани250"],
}

# Словник синонімів точок — часткові збіги
POINT_ALIASES = {
    "мурчик":        ["мурч", "мурчик", "мурчика"],
    "вдов":          ["вдов"],
    "сухопара":      ["сухоп", "сухопара"],
    "гордівка":      ["гордів", "гордівка"],
    "торканівка":    ["торкан", "торканівка"],
    "оляниця":       ["олян", "оляниця"],
    "мамина вишня":  ["мамин", "вишня", "мамина", "мамина вишня"],
    "клуб":          ["клуб"],
    "ася":           ["ася"],
    "ободівка агро": ["ободів", "ободівка", "агро"],
    "ковалівка":     ["ковалів", "ковалівка"],
    "корпуса":       ["корпус", "корпуса"],
}

def normalize_good(word: str) -> str | None:
    """Повертає стандартну назву товару або None"""
    w = word.lower().strip()
    for good, aliases in GOOD_ALIASES.items():
        if any(w.startswith(a) or a.startswith(w[:4]) for a in aliases if len(w) >= 3):
            return good
    return None

def normalize_point(text: str, points: dict) -> tuple[str, str] | tuple[None, None]:
    """Шукає назву точки в тексті"""
    tl = text.lower()
    # Спочатку точний збіг
    for point, owner in points.items():
        if point.lower() in tl:
            return point, owner
    # Потім по аліасах
    for key, aliases in POINT_ALIASES.items():
        for alias in aliases:
            if alias in tl:
                # Знаходимо реальну точку в словнику
                for point, owner in points.items():
                    if point.lower().startswith(key[:4]):
                        return point, owner
    return None, None

def smart_parse_supply(text: str) -> tuple[dict, float]:
    """Парсить постачання з вільного тексту"""
    items  = {}
    amount = 0.0
    tl     = text.lower()

    # Шукаємо паттерн: число + товар АБО товар + число
    # Наприклад: "20 комплектів", "+10 кав", "кава 5"
    tokens = re.findall(r'[\+\-]?\d+|[а-яіїєa-z]+\.?', tl)

    i = 0
    while i < len(tokens):
        tok = tokens[i]
        # Якщо це число
        if re.match(r'[\+\-]?\d+', tok):
            num = abs(int(tok))
            # Дивимось наступний токен — чи це товар?
            if i + 1 < len(tokens):
                good = normalize_good(tokens[i+1])
                if good:
                    items[good] = items.get(good, 0) + num
                    i += 2
                    continue
        # Якщо це товар — дивимось попередній або наступний токен
        else:
            good = normalize_good(tok)
            if good:
                # Шукаємо число поруч
                num = None
                if i > 0 and re.match(r'[\+\-]?\d+', tokens[i-1]):
                    num = abs(int(tokens[i-1]))
                elif i + 1 < len(tokens) and re.match(r'[\+\-]?\d+', tokens[i+1]):
                    num = abs(int(tokens[i+1]))
                    i += 1
                if num:
                    items[good] = items.get(good, 0) + num
        i += 1

    # Сума в грн
    m = re.search(r'(\d+)\s*грн', tl)
    if m:
        amount = float(m.group(1))

    return items, amount

async def smart_parse_free(text: str, d: dict, update, ctx, uid: int):
    """
    Локальний розумний парсинг вільного тексту.
    Порядок перевірок:
      1. Постачання (+ або ключові слова) — найвищий пріоритет
      2. Виплата
      3. Оренда / Обладнання
      4. Продаж (точка + товар)
      5. Тільки товар (питаємо точку)
    """
    tl = text.lower().strip()

    # ── 1. ПОСТАЧАННЯ — перевіряємо ПЕРШИМ ──
    # Ключові слова постачання (всі варіанти)
    supply_keywords = [
        "привіз", "поставив", "постачання", "прийшло", "привезли",
        "отримав", "закупка", "привіз коля", "коля привіз", "привіз товар",
        "прихід", "приход", "прийшов товар", "завіз", "закупили", "купили",
        "доставка товару", "товар прийшов",
    ]
    has_plus        = bool(re.search(r'\+\d', tl))
    has_supply_word = any(kw in tl for kw in supply_keywords)

    if has_plus or has_supply_word:
        items, amount = smart_parse_supply(text)
        if items:
            result = await do_supply(items, amount, d)
            await update.message.reply_text(result, parse_mode="Markdown")
            return result
        # Є ключове слово але не зрозуміли товар — питаємо
        if has_supply_word:
            QUICK[uid] = {"step": "supply_text"}
            await update.message.reply_text(
                "📥 Що саме привезли?\n_Наприклад: 20 комплектів 10 кава_",
                parse_mode="Markdown")
            return ""

    # ── 2. ВИПЛАТА ──
    pay_keywords = ["передав", "відав", "віддав", "оплатив", "заплатив",
                    "розрахував", "повернув", "зп", "зарплата"]
    m_amount = re.search(r'(\d+)', tl)
    if any(kw in tl for kw in pay_keywords) and m_amount:
        amount = float(m_amount.group())
        payer  = "vygran" if any(w in tl for w in ["коля","він","партнер"]) else "volodymyr"
        result = await do_payment(amount, payer, d, text)
        await update.message.reply_text(result, parse_mode="Markdown")
        return result

    # ── 3. ОРЕНДА ──
    if any(w in tl for w in ["оренда", "оренди", "оплата оренди"]):
        m = re.search(r'(\d+)', tl)
        if m:
            payer  = "vygran" if any(w in tl for w in ["коля","він"]) else "volodymyr"
            result = await do_expense("rent", float(m.group(1)), payer, d, text)
            await update.message.reply_text(result, parse_mode="Markdown")
            return result

    # ── 4. ОБЛАДНАННЯ ──
    if any(w in tl for w in ["купюрник", "принтер", "обладнання", "купив апарат"]):
        m = re.search(r'(\d+)', tl)
        if m:
            payer  = "vygran" if any(w in tl for w in ["коля","він"]) else "volodymyr"
            result = await do_expense("equipment", float(m.group(1)), payer, d, text)
            await update.message.reply_text(result, parse_mode="Markdown")
            return result

    # ── 5. ПРОДАЖ: точка + товар ──
    point, owner = normalize_point(tl, d["points"])
    if point:
        items  = {}
        tokens = re.findall(r'\d+|[а-яіїєa-z]+\.?', tl)
        i = 0
        while i < len(tokens):
            tok = tokens[i]
            if re.match(r'\d+', tok):
                num = int(tok)
                if i + 1 < len(tokens):
                    good = normalize_good(tokens[i+1])
                    if good:
                        items[good] = items.get(good, 0) + num
                        i += 2; continue
            else:
                good = normalize_good(tok)
                if good:
                    num = None
                    if i > 0 and re.match(r'\d+', tokens[i-1]):
                        num = int(tokens[i-1])
                    elif i+1 < len(tokens) and re.match(r'\d+', tokens[i+1]):
                        num = int(tokens[i+1]); i += 1
                    if num:
                        items[good] = items.get(good, 0) + num
                    else:
                        QUICK[uid] = {"step":"qty","point":point,"owner":owner,"good":good}
                        await update.message.reply_text(
                            f"📍 *{point}* — скільки *{good}*?",
                            parse_mode="Markdown")
                        return ""
            i += 1

        if items:
            result = await do_sale(point, owner, items, d)
            sale_kb = InlineKeyboardMarkup([[
                InlineKeyboardButton("🧾 Чек", callback_data="show_receipt"),
                InlineKeyboardButton("📋 У борг", callback_data="receipt_debt"),
            ]])
            await update.message.reply_text(result, parse_mode="Markdown", reply_markup=sale_kb)
            return result

        kb = InlineKeyboardMarkup([
            [InlineKeyboardButton("☕ Кава",      callback_data=f"pt2_{point}_кава"),
             InlineKeyboardButton("📦 Комплект", callback_data=f"pt2_{point}_комплект")],
            [InlineKeyboardButton("🥛 Молоко",   callback_data=f"pt2_{point}_молоко"),
             InlineKeyboardButton("🍹 Айріш",    callback_data=f"pt2_{point}_айріш")],
        ])
        await update.message.reply_text(
            f"📍 *{point}* — що продав?", parse_mode="Markdown", reply_markup=kb)
        return ""

    # ── 6. ТІЛЬКИ ТОВАР — питаємо точку ──
    solo_good = None
    solo_qty  = None
    tokens = re.findall(r'\d+|[а-яіїєa-z]+\.?', tl)
    for i, tok in enumerate(tokens):
        g = normalize_good(tok)
        if g:
            solo_good = g
            if i > 0 and re.match(r'\d+', tokens[i-1]):
                solo_qty = int(tokens[i-1])
            elif i+1 < len(tokens) and re.match(r'\d+', tokens[i+1]):
                solo_qty = int(tokens[i+1])
            break

    if solo_good:
        if solo_qty:
            QUICK[uid] = {"step":"point","good":solo_good,"pending_qty":solo_qty}
        else:
            QUICK[uid] = {"step":"point","good":solo_good}
        pts = list(d["points"].keys())
        kb  = InlineKeyboardMarkup([[InlineKeyboardButton(p, callback_data=f"pt_{p}")] for p in pts])
        msg = f"📍 {solo_qty} {solo_good} — обери точку продажу:" if solo_qty else f"📍 *{solo_good}* — обери точку:"
        await update.message.reply_text(msg, parse_mode="Markdown", reply_markup=kb)
        return ""

    return None  # Передаємо AI



async def on_text(update: Update, ctx: ContextTypes.DEFAULT_TYPE):
    uid  = update.effective_user.id
    if not ok(uid): return
    text = update.message.text.strip()
    d    = load()

    # PIN
    if uid in PIN_WAIT:
        mode = PIN_WAIT.pop(uid)
        if mode == "set":
            if re.match(r"^\d{4}$", text):
                pins = load_pins(); pins[str(uid)] = text; save_pins(pins)
                await update.message.reply_text("✅ PIN встановлено!")
            else:
                await update.message.reply_text("❌ PIN — 4 цифри.")
        elif mode == "check":
            pins = load_pins()
            if pins.get(str(uid)) == text:
                await show_balance(update)
            else:
                await update.message.reply_text("❌ Невірний PIN.")
        return

    # Кнопки меню
    if text == "💰 Баланс":        await cmd_balance(update, ctx); return
    if text == "📊 Звіт":          await cmd_report(update, ctx); return
    if text == "↩️ Скасувати останнє": await cmd_undo(update, ctx); return
    if text == "⚙️ Налаштування":  await cmd_settings(update, ctx); return
    if text == "📍 Точки":         await cmd_points(update, ctx); return
    if text == "🤝 Розрахунок з Колею": await cmd_settlement(update, ctx); return
    if text == "📊 Графік":              await cmd_chart(update, ctx); return
    if text == "🔔 Нагадати клієнтам":   await cmd_remind_clients(update, ctx); return
    if text == "👥 Борги клієнтів":     await cmd_clients(update, ctx); return
    if text == "💲 Змінити ціни":       await cmd_setprice(update, ctx); return
    if text == "🧾 Чек клієнту":        await cmd_receipt(update, ctx); return
    if text == "↩️ Скасувати":          await cmd_undo(update, ctx); return

    if text in ("☕ Продаж кави","📦 Продаж комплекту","🥛 Молоко / Айріш"):
        gmap = {"☕ Продаж кави":"кава","📦 Продаж комплекту":"комплект","🥛 Молоко / Айріш":"молоко/айріш"}
        QUICK[uid] = {"step":"point","good":gmap[text]}
        pts = list(d["points"].keys())
        kb  = InlineKeyboardMarkup([[InlineKeyboardButton(p, callback_data=f"pt_{p}")] for p in pts])
        await update.message.reply_text(f"📍 Обери точку:", reply_markup=kb)
        return

    if text == "📥 Постачання":
        QUICK[uid] = {"step":"supply_text"}
        await update.message.reply_text(
            "📥 Опиши постачання, наприклад:\n_+20 комплектів +10 кава 1500грн_",
            parse_mode="Markdown"); return

    if text == "💸 Виплата":
        kb = InlineKeyboardMarkup([[
            InlineKeyboardButton("Я → Колі",    callback_data="pay_vol"),
            InlineKeyboardButton("Коля → Мені", callback_data="pay_vyg"),
        ]])
        await update.message.reply_text("💸 Хто кому?", reply_markup=kb); return

    if text == "🏠 Оренда / Витрати":
        kb = InlineKeyboardMarkup([
            [InlineKeyboardButton("🏠 Оренда",      callback_data="ex_rent"),
             InlineKeyboardButton("🔧 Обладнання",  callback_data="ex_equipment")],
            [InlineKeyboardButton("🚚 Доставка",    callback_data="ex_delivery")],
        ])
        await update.message.reply_text("Тип витрати:", reply_markup=kb); return

    # Кроки діалогу
    state = QUICK.get(uid, {})
    step  = state.get("step","")

    if step == "qty":
        m = re.search(r"\d+", text)
        if not m: await update.message.reply_text("❌ Введи число."); return
        qty   = int(m.group())
        point = state["point"]; owner = state["owner"]; good = state["good"]
        QUICK.pop(uid, None)
        result = await do_sale(point, owner, {good: qty}, d)
        await update.message.reply_text(result, parse_mode="Markdown")
        await check_debt(ctx, d); return

    if step == "pay_amount":
        m = re.search(r"[\d.]+", text)
        if not m: await update.message.reply_text("❌ Введи суму."); return
        amount = float(m.group())
        payer  = state["payer"]
        QUICK.pop(uid, None)
        if amount >= CONFIRM_SUM:
            await ask_confirm(update, ctx, f"pay_{uid}_{int(amount)}",
                              f"Виплата {uname(ID_VOLODYMYR if payer=='volodymyr' else ID_VYGRAN)}",
                              amount, {"op":"payment","payer":payer,"comment":""})
            return
        else:
            result = await do_payment(amount, payer, d)
            await update.message.reply_text(result, parse_mode="Markdown")
            await check_debt(ctx, d); return

    if step == "exp_amount":
        m = re.search(r"[\d.]+", text)
        if not m: await update.message.reply_text("❌ Введи суму."); return
        amount = float(m.group())
        QUICK.pop(uid, None)
        result = await do_expense(state["etype"], amount, state["payer"], d)
        await update.message.reply_text(result, parse_mode="Markdown")
        await check_debt(ctx, d); return

    if step == "setprice_good":
        QUICK.pop(uid, None)
        good = state.get("good","кава")
        d2   = load()
        now  = datetime.now().strftime("%d.%m.%Y %H:%M")
        if good == "молоко":
            m = re.search(r"(\d+)", text)
            if not m: await update.message.reply_text("❌ Введи число грн"); return
            commission = int(m.group(1))
            d2.setdefault("prices",{}).setdefault("молоко",{})["his"]    = commission
            d2.setdefault("prices",{}).setdefault("айріш",{})["his"]     = commission
            d2.setdefault("price_history",[]).append({"date":now,"good":"молоко/айріш","commission":commission})
            save(d2)
            await update.message.reply_text(f"✅ Комісія молоко/айріш: *{commission} грн*", parse_mode="Markdown"); return
        parts = re.findall(r"\d+", text)
        if len(parts) < 2:
            await update.message.reply_text("❌ Введи два числа: _закупка продаж_\nНаприклад: _580 750_", parse_mode="Markdown"); return
        buy, sell = int(parts[0]), int(parts[1])
        commission = d2.get("prices",{}).get(good,{}).get("his", sell - 70)
        d2.setdefault("prices",{}).setdefault(good,{}).update({
            "buy": buy, "sell": sell,
            "my":  buy,
            "his": sell - 70,
        })
        d2.setdefault("price_history",[]).append({"date":now,"good":good,"buy":buy,"sell":sell})
        save(d2)
        await update.message.reply_text(
            f"✅ *{good}*: закупка {buy} грн / продаж {sell} грн\n"
            f"Комісія Колі: {sell-70} грн (фіксована -70)",
            parse_mode="Markdown"); return

    if step == "newclient_name":
        QUICK.pop(uid, None)
        if text.strip() == ".": await update.message.reply_text("❌ Скасовано."); return
        name = text.strip().title()
        d2   = load()
        d2.setdefault("clients",{})[name] = {"debt":0,"history":[]}
        save(d2)
        kb = InlineKeyboardMarkup([[
            InlineKeyboardButton("➕ Борг у борг",  callback_data=f"cl_debt_{name}"),
            InlineKeyboardButton("👤 Переглянути",  callback_data=f"cl_view_{name}"),
        ]])
        await update.message.reply_text(f"✅ Клієнт *{name}* доданий!", parse_mode="Markdown", reply_markup=kb); return

    if step == "debt_client_name":
        QUICK.pop(uid, None)
        name  = text.strip().title()
        d2    = load()
        sale  = state.get("sale", {})
        items = sale.get("items", {})
        prices = get_prices(d2)
        total  = sum(prices.get(g,{}).get("sell",0)*q for g,q in items.items())
        now    = datetime.now()
        c = d2.setdefault("clients",{}).setdefault(name, {"debt":0,"history":[]})
        c["debt"] = c.get("debt",0) + total
        c["last_date"] = now.isoformat()
        c["history"].append({"type":"debt","amount":total,"date":now.strftime("%d.%m.%Y"),"note":"продаж у борг"})
        save(d2)
        await update.message.reply_text(
            f"✅ Борг *{name}*: +{fm(total)} грн\nЗагальний: {fm(c['debt'])}",
            parse_mode="Markdown"); return

    if step == "client_pay":
        QUICK.pop(uid, None)
        name   = state.get("client","")
        m      = re.search(r"[\d.]+", text)
        if not m: await update.message.reply_text("❌ Введи суму"); return
        amount = float(m.group())
        d2     = load()
        c      = d2.get("clients",{}).get(name)
        if not c: await update.message.reply_text("❌ Клієнт не знайдений"); return
        c["debt"] = max(0, c.get("debt",0) - amount)
        c["last_date"] = datetime.now().isoformat()
        c.setdefault("history",[]).append({
            "type":"payment","amount":amount,
            "date":datetime.now().strftime("%d.%m.%Y"),"note":"оплата"})
        save(d2)
        await update.message.reply_text(
            f"✅ *{name}* заплатив {fm(amount)}\nЗалишок боргу: *{fm(c['debt'])}*",
            parse_mode="Markdown"); return

    if step == "client_add_debt":
        QUICK.pop(uid, None)
        name  = state.get("client","")
        m     = re.search(r"[\d.]+", text)
        if not m: await update.message.reply_text("❌ Введи суму"); return
        amount = float(m.group())
        note   = re.sub(r"[\d.]+", "", text).strip() or "борг"
        d2     = load()
        c      = d2.get("clients",{}).get(name)
        if not c: await update.message.reply_text("❌ Клієнт не знайдений"); return
        c["debt"] = c.get("debt",0) + amount
        c["last_date"] = datetime.now().isoformat()
        c.setdefault("history",[]).append({
            "type":"debt","amount":amount,
            "date":datetime.now().strftime("%d.%m.%Y"),"note":note})
        save(d2)
        await update.message.reply_text(
            f"✅ *{name}* — борг +{fm(amount)} ({note})\nЗагальний: *{fm(c['debt'])}*",
            parse_mode="Markdown"); return

    if step == "newpoint_name":
        QUICK.pop(uid, None)
        if text.strip() == ".":
            await update.message.reply_text("❌ Скасовано."); return
        name  = text.strip().title()
        owner = state.get("owner", "volodymyr")
        d["points"][name] = owner
        save(d)
        who = "моя" if owner == "volodymyr" else "Коли"
        kb = InlineKeyboardMarkup([[
            InlineKeyboardButton("➕ Ще одну мою",   callback_data="padd_vol"),
            InlineKeyboardButton("➕ Ще одну Коли",  callback_data="padd_vyg"),
        ]])
        await update.message.reply_text(
            f"✅ Точка *{name}* додана ({who})\n\nДодати ще?",
            parse_mode="Markdown", reply_markup=kb)
        return

    if step == "rename_point":
        QUICK.pop(uid, None)
        if text.strip() == ".":
            await update.message.reply_text("❌ Скасовано."); return
        old_name = state.get("old_name","")
        new_name = text.strip().title()
        if old_name in d["points"]:
            owner = d["points"].pop(old_name)
            d["points"][new_name] = owner
            save(d)
            who = "моя" if owner == "volodymyr" else "Коли"
            await update.message.reply_text(
                f"✅ *{old_name}* → *{new_name}* ({who})",
                parse_mode="Markdown")
        return

    if step == "supply_text":
        QUICK.pop(uid, None)
        items, amount = smart_parse_supply(text)
        if items:
            result = await do_supply(items, amount, d)
            await update.message.reply_text(result, parse_mode="Markdown")
            await check_debt(ctx, d)
        else:
            await update.message.reply_text(
                "❓ Не зрозумів — скористайся кнопками 👇",
                reply_markup=main_kb())
        return

    # Вільний текст — спочатку локальний парсинг, потім AI
    result = await smart_parse_free(text, d, update, ctx, uid)
    if result is not None:
        if result:
            save(d)
            await check_debt(ctx, d)
        return

    # AI як запасний варіант
    msg = await update.message.reply_text("🔍 Розпізнаю через AI...")
    parsed = await ai_parse(text)

    if not parsed or parsed.get("type") == "unknown":
        await msg.delete()
        await update.message.reply_text(
            "❓ Не зрозумів — скористайся кнопками 👇",
            reply_markup=main_kb()
        ); return

    ptype  = parsed.get("type","")
    items  = parsed.get("items") or {}
    amount = float(parsed.get("amount") or 0)
    point  = parsed.get("point") or ""
    payer  = parsed.get("payer") or "volodymyr"
    comment= parsed.get("comment") or ""
    result_text = ""

    if ptype == "sale" and point:
        pname, owner = find_point(point, d["points"])
        if pname:
            result_text = await do_sale(pname, owner, items, d)
        else:
            result_text = f"❓ Точка «{point}» не знайдена. /addpoint для додавання."
    elif ptype == "supply" and items:
        result_text = await do_supply(items, amount, d, comment)
    elif ptype == "payment" and amount:
        if amount >= CONFIRM_SUM:
            await msg.edit_text("⚠️ Велика сума — чекаю підтвердження...")
            await ask_confirm(update, ctx, f"pay_{uid}_{int(amount)}",
                              f"Виплата: {fm(amount)}", amount,
                              {"op":"payment","payer":payer,"comment":comment})
            return
        result_text = await do_payment(amount, payer, d, comment)
    elif ptype in ("rent","equipment","delivery") and amount:
        result_text = await do_expense(ptype, amount, payer, d, comment)
    else:
        result_text = "🤔 Не вистачає даних. Скористайся кнопками 👇"

    await msg.edit_text(result_text or "✅ Готово", parse_mode="Markdown")
    if result_text:
        save(d)
        await check_debt(ctx, d)

# ══════════════════════════════════════════════════════
# ГОЛОСОВІ ПОВІДОМЛЕННЯ
# ══════════════════════════════════════════════════════
async def on_voice(update: Update, ctx: ContextTypes.DEFAULT_TYPE):
    uid = update.effective_user.id
    if not ok(uid): return
    msg  = await update.message.reply_text("🎤 Транскрибую...")
    voice = update.message.voice
    file  = await ctx.bot.get_file(voice.file_id)
    with tempfile.NamedTemporaryFile(suffix=".ogg", delete=False) as tmp:
        await file.download_to_drive(tmp.name)
        text = await transcribe(tmp.name)
        os.unlink(tmp.name)
    if not text:
        await msg.edit_text("❌ Не вдалось розпізнати. Перевір GROQ_API_KEY.")
        return
    await msg.edit_text(f"🎤 *Розпізнав:* _{text}_", parse_mode="Markdown")
    update.message.text = text
    await on_text(update, ctx)

# ══════════════════════════════════════════════════════
# АВТО-ЗАДАЧІ
# ══════════════════════════════════════════════════════
# ══════════════════════════════════════════════════════
# ЗМІНА ЦІН
# ══════════════════════════════════════════════════════
async def cmd_setprice(update: Update, ctx: ContextTypes.DEFAULT_TYPE):
    uid = update.effective_user.id
    if not ok(uid): return
    d = load()
    p = get_prices(d)
    text = (
        "💲 *ПОТОЧНІ ЦІНИ*\n\n"
        f"☕ Кава: закупка *{p['кава']['buy']}* / продаж *{p['кава']['sell']}* грн\n"
        f"📦 Комплект: закупка *{p['комплект']['buy']}* / продаж *{p['комплект']['sell']}* грн\n\n"
        "Обери що змінити:"
    )
    kb = InlineKeyboardMarkup([
        [InlineKeyboardButton("☕ Ціна кави",      callback_data="sp_кава"),
         InlineKeyboardButton("📦 Ціна комплекту", callback_data="sp_комплект")],
        [InlineKeyboardButton("🥛 Молоко/Айріш",  callback_data="sp_молоко")],
    ])
    await update.message.reply_text(text, parse_mode="Markdown", reply_markup=kb)

# ══════════════════════════════════════════════════════
# ЧЕК КЛІЄНТУ
# ══════════════════════════════════════════════════════
async def cmd_receipt(update: Update, ctx: ContextTypes.DEFAULT_TYPE):
    uid = update.effective_user.id
    if not ok(uid): return
    d = load()
    receipt = d.get("_last_receipt")
    sale    = d.get("_last_sale", {})
    if not receipt:
        await update.message.reply_text("❌ Немає останнього продажу для чеку.")
        return
    kb = InlineKeyboardMarkup([
        [InlineKeyboardButton("📋 Чек у борг клієнту", callback_data="receipt_debt")],
        [InlineKeyboardButton("✅ Закрити",             callback_data="receipt_close")],
    ])
    await update.message.reply_text(
        f"```\n{receipt}\n```\n_Скопіюй і надішли клієнту_",
        parse_mode="Markdown", reply_markup=kb)

# ══════════════════════════════════════════════════════
# БОРГИ КЛІЄНТІВ
# ══════════════════════════════════════════════════════
async def cmd_clients(update: Update, ctx: ContextTypes.DEFAULT_TYPE):
    uid = update.effective_user.id
    if not ok(uid): return
    d       = load()
    clients = d.get("clients", {})

    if not clients:
        kb = InlineKeyboardMarkup([[
            InlineKeyboardButton("➕ Додати клієнта", callback_data="cl_add")
        ]])
        await update.message.reply_text(
            "👥 *Клієнтів поки немає*\nДодай першого:", 
            parse_mode="Markdown", reply_markup=kb)
        return

    # Підсумок боргів
    total_debt = sum(c.get("debt", 0) for c in clients.values())
    debtors    = [(n, c) for n, c in clients.items() if c.get("debt", 0) > 0]
    clean      = [(n, c) for n, c in clients.items() if c.get("debt", 0) <= 0]

    lines = ["👥 *БОРГИ КЛІЄНТІВ*\n"]
    if debtors:
        lines.append("🔴 *Борги:*")
        for name, c in sorted(debtors, key=lambda x: -x[1]["debt"]):
            days = (datetime.now() - datetime.fromisoformat(c.get("last_date", datetime.now().isoformat()))).days
            lines.append(f"  • {name}: *{fm(c['debt'])}* ({days}д тому)")
    if clean:
        lines.append("\n✅ *Без боргів:*")
        for name, c in clean:
            lines.append(f"  • {name}")
    lines.append(f"\n💰 *Загальний борг: {fm(total_debt)}*")

    kb_rows = []
    for name in clients:
        short = name[:14]+"…" if len(name) > 14 else name
        debt  = clients[name].get("debt", 0)
        label = f"{short} ({fm(debt)})" if debt > 0 else short
        kb_rows.append([
            InlineKeyboardButton(f"👤 {label}", callback_data=f"cl_view_{name}"),
        ])
    kb_rows.append([InlineKeyboardButton("➕ Новий клієнт", callback_data="cl_add")])
    await update.message.reply_text(
        "\n".join(lines), parse_mode="Markdown",
        reply_markup=InlineKeyboardMarkup(kb_rows))

# ══════════════════════════════════════════════════════
# РОЗРАХУНОК З КОЛЕЮ

# ══════════════════════════════════════════════════════
def build_settlement_text(d: dict) -> str:
    """Формує повний звіт для розрахунку"""
    b  = d.get("balance", 0)
    s  = d.get("stock", {})
    sm = d.get("stats", {}).get("my", {})
    sh = d.get("stats", {}).get("his", {})
    p  = d.get("prices", {})

    # Вартість залишків (за собівартістю)
    stock_val = (
        s.get("кава", 0)     * p.get("kava_buy", 530) +
        s.get("комплект", 0) * p.get("komplex_buy", 680)
    )

    # Прибуток Володимира
    profit = (
        sm.get("кава", 0)     * (p.get("kava_sell", 700) - p.get("kava_buy", 530)) +
        sm.get("комплект", 0) * (p.get("komplex_sell", 850) - p.get("komplex_buy", 680)) +
        sh.get("кава", 0)     * p.get("commission", 70) +
        sh.get("комплект", 0) * p.get("commission", 70) +
        (sh.get("молоко", 0) + sh.get("айріш", 0)) * p.get("milk_commission", 100)
    )

    lines = [
        "🤝 *РОЗРАХУНОК З КОЛЕЮ*",
        f"{'═'*28}",
        "",
        "💰 *БАЛАНС:*",
    ]

    if abs(b) < 1:
        lines.append("  ✅ Рахунки зведені — нікому нічого не винні")
    elif b > 0:
        lines.append(f"  🔴 Коля винен Володимиру: *{fm(b)}*")
    else:
        lines.append(f"  🔴 Володимир винен Колі: *{fm(abs(b))}*")

    lines += [
        "",
        "📦 *ЗАЛИШКИ НА СКЛАДІ:*",
    ]

    em = {"кава": "☕", "комплект": "📦", "молоко": "🥛",
          "айріш": "🍹", "стакан110": "🥤", "стакан250": "🥤"}
    has_stock = False
    for good, qty in s.items():
        if qty > 0:
            has_stock = True
            lines.append(f"  {em.get(good,'•')} {good}: *{qty} шт*")
    if not has_stock:
        lines.append("  Склад порожній")
    if stock_val > 0:
        lines.append(f"  💵 Вартість залишків: *{fm(stock_val)}*")

    lines += [
        "",
        "📊 *ПРОДАНО ЗА ВЕСЬ ПЕРІОД:*",
        f"  Мої точки: кава {sm.get('кава',0)} | компл {sm.get('комплект',0)}",
        f"  Точки Колі: кава {sh.get('кава',0)} | компл {sh.get('комплект',0)} "
        f"| мол/айр {sh.get('молоко',0)+sh.get('айріш',0)}",
        "",
        f"{'─'*28}",
        f"💚 *Прибуток Володимира: {fm(profit)}*",
        f"{'═'*28}",
    ]

    # Висновок — хто платить
    if abs(b) < 1:
        lines.append("✅ *Нікому платити не треба!*")
    elif b > 0:
        lines.append(f"👉 *Коля платить Володимиру: {fm(b)}*")
    else:
        lines.append(f"👉 *Володимир платить Колі: {fm(abs(b))}*")

    return "\n".join(lines)

async def cmd_settlement(update: Update, ctx: ContextTypes.DEFAULT_TYPE):
    uid = update.effective_user.id
    if not ok(uid): return
    d    = load()
    text = build_settlement_text(d)

    kb = InlineKeyboardMarkup([
        [InlineKeyboardButton("📤 Надіслати Колі",      callback_data="settle_send")],
        [InlineKeyboardButton("✅ Оплачено — обнулити баланс", callback_data="settle_paid")],
        [InlineKeyboardButton("❌ Закрити",             callback_data="settle_close")],
    ])
    await update.message.reply_text(text, parse_mode="Markdown", reply_markup=kb)

async def check_debt(ctx: ContextTypes.DEFAULT_TYPE, d: dict):
    b = abs(d.get("balance", 0))
    if b >= DEBT_ALERT:
        debtor = "Коля" if d["balance"] > 0 else "Володимир"
        cred   = "Володимиру" if d["balance"] > 0 else "Колі"
        msg    = f"🔔 *Нагадування!*\n{debtor} винен {cred} *{fm(b)}*\nЧас розрахуватись? 😊"
        for u in (ID_VOLODYMYR, ID_VYGRAN):
            try: await ctx.bot.send_message(u, msg, parse_mode="Markdown")
            except: pass

async def weekly_job(ctx: ContextTypes.DEFAULT_TYPE):
    if datetime.now().weekday() != 0: return   # тільки понеділок
    d     = load()
    today = datetime.now().strftime("%Y-%m-%d")
    if d.get("weekly_sent") == today: return
    txs   = d.get("transactions",[])
    week  = (datetime.now()-timedelta(days=7)).strftime("%Y-%m-%d")
    wtxs  = [t for t in txs if t["date"][:10] >= week]
    sales = [t for t in wtxs if t["type"]=="sale"]
    msg   = (f"📊 *Тижневий звіт*\n\n"
             f"Продажів за тиждень: {len(sales)}\n"
             f"Сума: {fm(sum(t['amount'] for t in sales))}\n\n"
             f"{bal_line(d['balance'])}")
    for u in (ID_VOLODYMYR, ID_VYGRAN):
        try: await ctx.bot.send_message(u, msg, parse_mode="Markdown")
        except: pass
    d["weekly_sent"] = today
    save(d)

# ══════════════════════════════════════════════════════
# ФОТО ЧЕКУ — AI розпізнає суму
# ══════════════════════════════════════════════════════
async def on_photo(update: Update, ctx: ContextTypes.DEFAULT_TYPE):
    uid = update.effective_user.id
    if not ok(uid): return
    if not GROQ_API_KEY or GROQ_API_KEY == "ВАШ_GROQ_KEY":
        await update.message.reply_text("❌ Groq API не налаштований")
        return

    msg = await update.message.reply_text("🔍 Розпізнаю чек через AI...")
    # Беремо найбільше фото
    photo = update.message.photo[-1]
    f     = await ctx.bot.get_file(photo.file_id)
    import tempfile, base64
    with tempfile.NamedTemporaryFile(suffix=".jpg", delete=False) as tmp:
        await f.download_to_drive(tmp.name)
        img_path = tmp.name

    try:
        with open(img_path, "rb") as img_f:
            b64 = base64.b64encode(img_f.read()).decode()
        os.unlink(img_path)

        headers = {"Authorization": f"Bearer {GROQ_API_KEY}", "Content-Type": "application/json"}
        payload = {
            "model": "meta-llama/llama-4-scout-17b-16e-instruct",
            "messages": [{
                "role": "user",
                "content": [
                    {"type": "image_url", "image_url": {"url": f"data:image/jpeg;base64,{b64}"}},
                    {"type": "text", "text": (
                        "Це фото чеку або накладної. Визнач: "
                        "1) загальну суму до оплати (тільки число), "
                        "2) дату якщо є, "
                        "3) перелік товарів якщо видно. "
                        'Відповідай JSON: {"total": 1234, "date": "01.01.2026", "items": "кава 5кг"}'
                    )}
                ]
            }],
            "max_tokens": 300,
        }
        async with aiohttp.ClientSession() as s:
            async with s.post("https://api.groq.com/openai/v1/chat/completions",
                              json=payload, headers=headers) as r:
                data = await r.json()

        raw = data["choices"][0]["message"]["content"]
        # Витягуємо JSON
        m = re.search(r'\{.*\}', raw, re.DOTALL)
        if m:
            info = json.loads(m.group())
            total = info.get("total", 0)
            date  = info.get("date", "")
            items = info.get("items", "")
            txt = "*Чек*" + chr(10) + f"Сума: *{fm(total)}*"

            if date:  txt += f"  Дата: {date}"
            if items: txt += f"  Товари: {items}"
            kb = InlineKeyboardMarkup([
                [InlineKeyboardButton(f"✅ Записати витрату {fm(total)}", callback_data=f"photo_exp_{int(total)}")],
                [InlineKeyboardButton(f"📥 Записати постачання",          callback_data=f"photo_sup_{int(total)}")],
                [InlineKeyboardButton("❌ Скасувати",                     callback_data="photo_cancel")],
            ])
            await msg.edit_text(txt, parse_mode="Markdown", reply_markup=kb)
        else:
            await msg.edit_text("AI не розпізнав. " + raw[:150])


    except Exception as e:
        await msg.edit_text(f"❌ Помилка розпізнавання: {e}")

# ══════════════════════════════════════════════════════
# ГРАФІК ПРОДАЖІВ
# ══════════════════════════════════════════════════════
async def cmd_chart(update, ctx):
    uid = update.effective_user.id
    if not ok(uid): return
    d = load()
    txs = [t for t in d.get("transactions",[]) if t["type"]=="sale"]
    if not txs: await update.message.reply_text("Немає продажів."); return
    from collections import defaultdict
    daily = defaultdict(float)
    cut = (datetime.now()-timedelta(days=13)).strftime("%Y-%m-%d")
    for t in txs:
        if t["date"][:10] >= cut: daily[t["date"][:10]] += t.get("amount",0)
    if not daily: await update.message.reply_text("Немає даних за 14 днів."); return
    days = sorted(daily.keys())
    vals = [daily[d] for d in days]
    mv   = max(vals) or 1
    rows = ["*Продажі за 14 днів*"]
    for day,val in zip(days,vals):
        n = int(val/mv*8)
        rows.append(f"`{day[5:]}` {chr(9608)*n}{chr(9617)*(8-n)} {int(val/1000)}к")
    rows.append(f"Макс: {fm(mv)}  Всього: {fm(sum(vals))}")
    await update.message.reply_text(chr(10).join(rows), parse_mode="Markdown")


async def cmd_remind_clients(update, ctx):
    uid = update.effective_user.id
    if not ok(uid): return
    d = load()
    clients = {n:c for n,c in d.get("clients",{}).items() if c.get("debt",0)>0}
    if not clients:
        await update.message.reply_text("Жодного боргу!"); return
    for name,c in sorted(clients.items(),key=lambda x:-x[1]["debt"]):
        debt = c["debt"]
        days = 0
        if c.get("last_date"):
            try: days=(datetime.now()-datetime.fromisoformat(c["last_date"])).days
            except: pass
        parts = ["Привіт, "+name+"!", "Борг за каву: "+str(int(debt))+" грн"]
        if days: parts.append("Брали "+str(days)+" днів тому.")
        parts.append("Розрахуйся при наступному замовленні. Дякуємо!")
        reminder = chr(10).join(parts)
        kb = InlineKeyboardMarkup([[InlineKeyboardButton("Оплачено", callback_data=f"cl_pay_{name}")]])
        await update.message.reply_text(
            f"*{name}* — {fm(debt)}"+chr(10)+"```"+chr(10)+reminder+chr(10)+"```",
            parse_mode="Markdown", reply_markup=kb)

async def daily_jobs(ctx):
    now = datetime.now()
    d   = load()
    nl  = chr(10)
    # Тижневий звіт (понеділок)
    if now.weekday() == 0:
        today = now.strftime("%Y-%m-%d")
        if d.get("weekly_sent") != today:
            txs   = d.get("transactions",[])
            week  = (now-timedelta(days=7)).strftime("%Y-%m-%d")
            sales = [t for t in txs if t["type"]=="sale" and t["date"][:10]>=week]
            msg   = ("*Тижневий звіт*"+nl
                     +f"Продажів: {len(sales)}"+nl
                     +f"Сума: {fm(sum(t["amount"] for t in sales))}"+nl
                     +bal_line(d["balance"]))
            await notify_all(ctx.bot, msg)
            d["weekly_sent"] = today
            save(d)
    # Нагадування про оренду
    if RENT_AMOUNT > 0 and now.day == RENT_DAY:
        month = now.strftime("%Y-%m")
        if d.get("rent_reminded") != month:
            await notify_all(ctx.bot,
                "*Нагадування про оренду*"+nl
                +f"Оренда: *{fm(RENT_AMOUNT)}*"+nl
                +"/rent — записати оплату")
            d["rent_reminded"] = month
            save(d)
    # Борги клієнтів (п'ятниця)
    if now.weekday() == 4:
        debtors = [(n,c["debt"]) for n,c in d.get("clients",{}).items() if c.get("debt",0)>0]
        if debtors:
            total = sum(v for _,v in debtors)
            rows  = ["*Борги клієнтів* (п'ятниця)"]
            for name,debt in sorted(debtors,key=lambda x:-x[1]):
                rows.append(f"  {name}: {fm(debt)}")
            rows.append(f"Всього: *{fm(total)}*")
            rows.append("/clients — переглянути")
            await notify_all(ctx.bot, nl.join(rows))

async def cmd_rent(update: Update, ctx: ContextTypes.DEFAULT_TYPE):
    """Швидка команда записати оренду"""
    uid = update.effective_user.id
    if not ok(uid): return
    d = load()
    args = ctx.args
    amount = RENT_AMOUNT
    if args:
        try: amount = int(args[0])
        except: pass
    if not amount:
        QUICK[uid] = {"step":"exp_amount","etype":"rent","payer":"volodymyr"}
        await update.message.reply_text("🏠 Введи суму оренди (грн):")
        return
    kb = InlineKeyboardMarkup([[
        InlineKeyboardButton("Я платив",      callback_data="ep_volodymyr"),
        InlineKeyboardButton("Коля платив",   callback_data="ep_vygran"),
    ]])
    QUICK[uid] = {"step":"exp_payer","etype":"rent","amount":amount}
    await update.message.reply_text(f"🏠 Оренда {fm(amount)} — хто платив?", reply_markup=kb)

# ══════════════════════════════════════════════════════
# ВЕБ-СЕРВЕР (aiohttp) — роздає дашборд і API
# ══════════════════════════════════════════════════════
from aiohttp import web as aioWeb
import asyncio

WEB_PORT = int(os.environ.get("PORT", 8080))

async def web_index(request):
    """Роздає dashboard.html"""
    html_path = os.path.join(os.path.dirname(__file__), "dashboard.html")
    if not os.path.exists(html_path):
        return aioWeb.Response(text="dashboard.html not found", status=404)
    with open(html_path, "r", encoding="utf-8") as f:
        content = f.read()
    return aioWeb.Response(text=content, content_type="text/html", charset="utf-8")

async def api_get_data(request):
    """API: повертає всі дані (для синхронізації з веб)"""
    # Простий захист — перевіряємо секретний токен в заголовку
    token = request.headers.get("X-Token", "")
    if token != BOT_TOKEN[:20]:
        return aioWeb.json_response({"error": "unauthorized"}, status=401)
    d = load()
    return aioWeb.json_response(d)

async def api_post_data(request):
    """API: зберігає дані з веб-дашборду"""
    token = request.headers.get("X-Token", "")
    if token != BOT_TOKEN[:20]:
        return aioWeb.json_response({"error": "unauthorized"}, status=401)
    try:
        body = await request.json()
        if "transactions" not in body:
            return aioWeb.json_response({"error": "invalid data"}, status=400)
        save(body)
        return aioWeb.json_response({"ok": True})
    except Exception as e:
        return aioWeb.json_response({"error": str(e)}, status=500)

async def api_health(request):
    """Health check для Railway"""
    return aioWeb.json_response({"status": "ok", "bot": "☕ КавоБот"})

def make_web_app():
    app = aioWeb.Application()
    app.router.add_get("/",        web_index)
    app.router.add_get("/health",  api_health)
    app.router.add_get("/api/data",  api_get_data)
    app.router.add_post("/api/data", api_post_data)
    return app

# ══════════════════════════════════════════════════════
# ЗАПУСК — БОТ + ВЕБ СЕРВЕР РАЗОМ
# ══════════════════════════════════════════════════════
async def run_bot(tg_app):
    """Запускає Telegram бота"""
    await tg_app.initialize()
    await tg_app.start()
    await tg_app.updater.start_polling(allowed_updates=["message","callback_query"])
    logger.info("☕ КавоБот запущено!")

async def run_web():
    """Запускає веб-сервер"""
    web_app = make_web_app()
    runner  = aioWeb.AppRunner(web_app)
    await runner.setup()
    site = aioWeb.TCPSite(runner, "0.0.0.0", WEB_PORT)
    await site.start()
    logger.info(f"🌐 Веб-сервер: http://0.0.0.0:{WEB_PORT}")

async def main_async():
    os.makedirs(EXPORT_DIR, exist_ok=True)
    # Завантажуємо групу з data.json якщо є
    global ID_GROUP_CHAT
    try:
        d0 = load()
        if d0.get("group_chat_id") and not ID_GROUP_CHAT:
            ID_GROUP_CHAT = d0["group_chat_id"]
            logger.info(f"Група: {ID_GROUP_CHAT}")
    except Exception:
        pass

    tg_app = Application.builder().token(BOT_TOKEN).build()
    tg_app.add_handler(CommandHandler("start",    cmd_start))
    tg_app.add_handler(CommandHandler("balance",  cmd_balance))
    tg_app.add_handler(CommandHandler("report",   cmd_report))
    tg_app.add_handler(CommandHandler("export",   cmd_export))
    tg_app.add_handler(CommandHandler("undo",     cmd_undo))
    tg_app.add_handler(CommandHandler("points",   cmd_points))
    tg_app.add_handler(CommandHandler("addpoint", cmd_addpoint))
    tg_app.add_handler(CommandHandler("stock",    cmd_stock))
    tg_app.add_handler(CommandHandler("history",  cmd_history))
    tg_app.add_handler(CommandHandler("setpin",   cmd_setpin))
    tg_app.add_handler(CommandHandler("settings",   cmd_settings))
    tg_app.add_handler(CommandHandler("rozrahunok", cmd_settlement))
    tg_app.add_handler(CommandHandler("clients",    cmd_clients))
    tg_app.add_handler(CommandHandler("setprice",   cmd_setprice))
    tg_app.add_handler(CommandHandler("receipt",    cmd_receipt))
    tg_app.add_handler(CommandHandler("setgroup",   cmd_setgroup))
    tg_app.add_handler(CommandHandler("chart",      cmd_chart))
    tg_app.add_handler(CommandHandler("remind",     cmd_remind_clients))
    tg_app.add_handler(CommandHandler("rent",       cmd_rent))
    tg_app.add_handler(MessageHandler(filters.TEXT & ~filters.COMMAND, on_text))
    tg_app.add_handler(MessageHandler(filters.VOICE, on_voice))
    tg_app.add_handler(MessageHandler(filters.PHOTO, on_photo))
    tg_app.add_handler(CallbackQueryHandler(on_callback))
    tg_app.job_queue.run_daily(
        daily_jobs,
        time=datetime.strptime("09:00", "%H:%M").time()
    )

    # Запускаємо веб і бота одночасно
    await run_web()
    await run_bot(tg_app)

    # Тримаємо живим
    try:
        await asyncio.Event().wait()
    except (KeyboardInterrupt, SystemExit):
        logger.info("Зупинка...")
        await tg_app.updater.stop()
        await tg_app.stop()
        await tg_app.shutdown()

def main():
    asyncio.run(main_async())

if __name__ == "__main__":
    main()
