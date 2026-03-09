"""
☕ КавоБот — облік кавового бізнесу Володимир + Коля
pip install python-telegram-bot[job-queue]==20.7 aiohttp openpyxl
"""
import json, os, re, logging, tempfile, base64, asyncio
from datetime import datetime, timedelta
from collections import defaultdict

from telegram import (Update, InlineKeyboardButton, InlineKeyboardMarkup,
                      ReplyKeyboardMarkup)
from telegram.ext import (Application, CommandHandler, MessageHandler,
                          CallbackQueryHandler, ContextTypes, filters)
import aiohttp

logging.basicConfig(format="%(asctime)s %(levelname)s %(message)s", level=logging.INFO)
logger = logging.getLogger(__name__)

BOT_TOKEN    = os.environ.get()
GROQ_API_KEY = os.environ.get("GROQ_API_KEY", "")

ID_VOLODYMYR  = int(os.environ.get("ID_VOLODYMYR",  "373296886"))
ID_VYGRAN     = int(os.environ.get("ID_VYGRAN",     "987654321"))
ID_GROUP_CHAT = int(os.environ.get("ID_GROUP_CHAT", "0"))

DEBT_ALERT   = 15000
CONFIRM_SUM  = 5000
LOW_STOCK    = int(os.environ.get("LOW_STOCK", "5"))  # поріг сповіщення про малий запас
RENT_DAY    = int(os.environ.get("RENT_DAY",    "1"))
RENT_AMOUNT = int(os.environ.get("RENT_AMOUNT", "0"))

DATA_FILE  = "data.json"
PIN_FILE   = "pins.json"
EXPORT_DIR = "exports"
QUICK: dict = {}
PENDING_CONFIRM: dict = {}

def default_data():
    return {
        "points": {}, "transactions": [],
        "stock":  {"кава":0,"комплект":0,"молоко":0,"айріш":0,"стакан110":0,"стакан250":0},
        "balance": 0.0,
        "stats":  {"my":{"кава":0,"комплект":0,"молоко":0,"айріш":0},
                   "his":{"кава":0,"комплект":0,"молоко":0,"айріш":0}},
        "clients": {}, "prices": {}, "price_history": [],
        "weekly_sent": "", "group_chat_id": 0, "quick_sales": [],
    }

def load():
    if os.path.exists(DATA_FILE):
        try:
            with open(DATA_FILE, encoding="utf-8") as f:
                d = json.load(f)
            for k,v in default_data().items():
                d.setdefault(k, v)
            return d
        except Exception:
            pass
    return default_data()

def save(d):
    with open(DATA_FILE, "w", encoding="utf-8") as f:
        json.dump(d, f, ensure_ascii=False, indent=2)

def add_tx(d, ttype, desc, amount, delta, meta=None):
    d["balance"] = round(d.get("balance",0) + delta, 2)
    d.setdefault("transactions",[]).append({
        "id": len(d["transactions"])+1, "date": datetime.now().isoformat(),
        "type": ttype, "desc": desc, "amount": round(abs(amount),2),
        "delta": round(delta,2), "balance": d["balance"], "meta": meta or {},
    })

def ok(uid):
    return uid in (ID_VOLODYMYR, ID_VYGRAN) or (ID_GROUP_CHAT and uid == ID_GROUP_CHAT)

def fm(v):
    return f"{abs(v):,.0f} грн".replace(",", " ")

def bal_line(b):
    if abs(b) < 1: return "✅ Рахунки зведені"
    if b > 0:      return f"🔴 Коля винен Тобі: *{fm(b)}*"
    return             f"🔴 Ти винен Колі: *{fm(abs(b))}*"

def get_prices(d):
    base = {
        "кава":      {"buy":530,"sell":700,"my":530,"his":630},
        "комплект":  {"buy":680,"sell":850,"my":680,"his":780},
        "молоко":    {"buy":0,  "sell":0,  "my":0,  "his":100},
        "айріш":     {"buy":0,  "sell":0,  "my":0,  "his":100},
        "стакан110": {"buy":0,  "sell":0,  "my":0,  "his":0},
        "стакан250": {"buy":0,  "sell":0,  "my":0,  "his":0},
    }
    for g,vals in d.get("prices",{}).items():
        if g in base: base[g].update(vals)
    return base

def uname(uid):
    return "Володимир" if uid == ID_VOLODYMYR else "Коля"

async def notify_all(bot, msg, parse_mode="Markdown", kb=None):
    global ID_GROUP_CHAT
    if not ID_GROUP_CHAT:
        try: ID_GROUP_CHAT = load().get("group_chat_id",0)
        except Exception: pass
    targets = [ID_GROUP_CHAT] if ID_GROUP_CHAT else [ID_VOLODYMYR, ID_VYGRAN]
    for t in targets:
        try: await bot.send_message(t, msg, parse_mode=parse_mode, reply_markup=kb)
        except Exception as e: logger.warning(f"notify {t}: {e}")

def main_kb():
    return ReplyKeyboardMarkup([
        ["⚡ Швидкий продаж",     "📦 Продаж комплекту"],
        ["☕ Продаж кави",        "🥛 Молоко / Айріш"],
        ["📥 Постачання",         "💸 Виплата"],
        ["🏠 Оренда / Витрати",   "💰 Баланс"],
        ["📊 Звіт",               "🏆 Топ точок"],
        ["📈 Графік",             "📋 Excel звіт"],
        ["📍 Точки",              "👥 Борги клієнтів"],
        ["💲 Змінити ціни",       "🔔 Нагадати клієнтам"],
        ["🤝 Розрахунок з Колею", "↩️ Скасувати"],
        ["⚙️ Налаштування"],
    ], resize_keyboard=True)

async def ai_parse(text):
    if not GROQ_API_KEY: return {}
    sys_p = 'Розпізнай операцію. JSON тільки: {"type":"sale","point":"назва","items":{"кава":5},"owner":"volodymyr"} або {"type":"supply","items":{},"amount":0} або {"type":"payment","amount":0,"payer":"volodymyr"} або {"type":"expense","etype":"rent","amount":0,"payer":"volodymyr"} або {"type":"unknown"}'
    try:
        async with aiohttp.ClientSession() as s:
            async with s.post("https://api.groq.com/openai/v1/chat/completions",
                headers={"Authorization":f"Bearer {GROQ_API_KEY}"},
                json={"model":"llama3-8b-8192","max_tokens":200,
                      "messages":[{"role":"system","content":sys_p},{"role":"user","content":text}]},
                timeout=aiohttp.ClientTimeout(total=8)) as r:
                data = await r.json()
        raw = data["choices"][0]["message"]["content"]
        m = re.search(r"\{.*\}", raw, re.DOTALL)
        return json.loads(m.group()) if m else {}
    except Exception as e:
        logger.warning(f"ai_parse: {e}"); return {}

async def ai_photo(b64):
    if not GROQ_API_KEY: return {}
    try:
        async with aiohttp.ClientSession() as s:
            async with s.post("https://api.groq.com/openai/v1/chat/completions",
                headers={"Authorization":f"Bearer {GROQ_API_KEY}"},
                json={"model":"meta-llama/llama-4-scout-17b-16e-instruct","max_tokens":200,
                      "messages":[{"role":"user","content":[
                          {"type":"image_url","image_url":{"url":f"data:image/jpeg;base64,{b64}"}},
                          {"type":"text","text":'JSON тільки: {"total":1234,"date":"01.01.2026","items":"назва"}'},
                      ]}]},
                timeout=aiohttp.ClientTimeout(total=15)) as r:
                data = await r.json()
        raw = data["choices"][0]["message"]["content"]
        m = re.search(r"\{.*\}", raw, re.DOTALL)
        return json.loads(m.group()) if m else {}
    except Exception as e:
        logger.warning(f"ai_photo: {e}"); return {}

async def transcribe(path):
    if not GROQ_API_KEY: return ""
    try:
        with open(path,"rb") as f:
            async with aiohttp.ClientSession() as s:
                form = aiohttp.FormData()
                form.add_field("file", f, filename="voice.ogg", content_type="audio/ogg")
                form.add_field("model","whisper-large-v3")
                form.add_field("language","uk")
                async with s.post("https://api.groq.com/openai/v1/audio/transcriptions",
                    headers={"Authorization":f"Bearer {GROQ_API_KEY}"},
                    data=form, timeout=aiohttp.ClientTimeout(total=20)) as r:
                    data = await r.json()
        return data.get("text","")
    except Exception as e:
        logger.warning(f"transcribe: {e}"); return ""

def make_receipt(point, items, prices):
    now = datetime.now().strftime("%d.%m.%Y %H:%M")
    lines = ["━━━━━━━━━━━━━━━━━━━━","       ☕ КавоМаркет",f"  {now}",f"  Точка: {point}","━━━━━━━━━━━━━━━━━━━━"]
    total = 0
    for good,qty in items.items():
        sell = prices.get(good,{}).get("sell",0)
        sub  = sell*qty; total += sub
        lines.append(f"  {good:<12} {qty}шт×{sell}={sub}грн")
    lines += ["━━━━━━━━━━━━━━━━━━━━",f"  РАЗОМ:  {total} грн","━━━━━━━━━━━━━━━━━━━━","  Дякуємо! ☕"]
    return "\n".join(lines)

async def do_sale(point, owner, items, d):
    prices = get_prices(d)
    total_pay = total_rev = 0
    lines = [f"☕ *Продаж — {point}* ({'моя' if owner=='volodymyr' else 'Колі'})\n"]
    for good,qty in items.items():
        p = prices.get(good)
        if not p: continue
        sell = p["sell"]; pay = p["my"] if owner=="volodymyr" else p["his"]
        rev = sell*qty; total_pay += pay*qty; total_rev += rev
        lines.append(f"  {good}: {qty}шт × {sell} = {fm(rev)} → Колі: {fm(pay*qty)}")
        if good in d["stock"]: d["stock"][good] = max(0, d["stock"][good]-qty)
        sk = "my" if owner=="volodymyr" else "his"
        d["stats"][sk][good] = d["stats"][sk].get(good,0)+qty
    earn = total_rev-total_pay
    lines += [f"\n💰 До виплати Колі: *{fm(total_pay)}*", f"💵 Виручка: {fm(total_rev)}", f"📈 Заробіток: *{fm(earn)}*"]
    add_tx(d,"sale",f"Продаж {point}: {items}",total_pay,-total_pay,{"point":point,"owner":owner,"items":items})
    for good,qty in items.items():
        rem = d["stock"].get(good,0)
        if rem <= LOW_STOCK: lines.append(f"⚠️ {good}: залишилось {rem} шт!")
    debtors = [(n,c["debt"]) for n,c in d.get("clients",{}).items() if c.get("debt",0)>0]
    if debtors:
        lines.append("\n👥 *Клієнти з боргами:*")
        for n,debt in sorted(debtors,key=lambda x:-x[1])[:3]:
            lines.append(f"  • {n}: {fm(debt)}")
    d["_last_receipt"] = make_receipt(point, items, prices)
    d["_last_sale"]    = {"point":point,"owner":owner,"items":items}
    save(d)
    return "\n".join(lines)

async def do_supply(items, amount, d, comment=""):
    prices = get_prices(d)
    total = 0
    lines = ["📥 *Постачання від Колі*\n"]
    for good,qty in items.items():
        if good in d["stock"]: d["stock"][good] += qty
        buy = prices.get(good,{}).get("buy",0); sub = buy*qty; total += sub
        lines.append(f"  {good}: +{qty} шт" + (f" = {fm(sub)}" if sub else ""))
        lines.append(f"  → Склад: {d['stock'].get(good,qty)} шт")
    if amount: total = amount
    if comment: lines.append(f"\n💬 {comment}")
    if total:
        lines.append(f"\n💰 Сума: *{fm(total)}*")
        lines.append(f"📊 Баланс: {bal_line(d['balance']+total)}")
    add_tx(d,"supply",f"Постачання: {items}",total,total,{"items":items,"comment":comment})
    save(d); return "\n".join(lines)

async def do_payment(amount, payer, d, comment=""):
    if payer=="volodymyr":
        delta=amount; desc=f"Виплата Колі: {fm(amount)}"; txt=f"💸 Ти заплатив Колі *{fm(amount)}*"
    else:
        delta=-amount; desc=f"Виплата від Колі: {fm(amount)}"; txt=f"💸 Коля заплатив Тобі *{fm(amount)}*"
    add_tx(d,"payment",desc,amount,delta,{"payer":payer,"comment":comment})
    save(d); return f"{txt}\n\n{bal_line(d['balance'])}"

async def do_expense(etype, amount, payer, d, comment=""):
    labels={"rent":"🏠 Оренда","equipment":"🔧 Обладнання","delivery":"🚚 Доставка"}
    label=labels.get(etype,"💸 Витрата")
    if payer=="volodymyr":
        delta=amount/2; txt=f"{label} {fm(amount)} — Ти платив\nКоля відшкодує *{fm(amount/2)}*"
    else:
        delta=-amount/2; txt=f"{label} {fm(amount)} — Коля платив\nТи відшкодуєш *{fm(amount/2)}*"
    add_tx(d,etype,f"{label}: {fm(amount)}",amount/2,delta,{"etype":etype,"payer":payer,"comment":comment})
    save(d); return f"{txt}\n\n{bal_line(d['balance'])}"

async def cmd_start(update, ctx):
    uid = update.effective_user.id
    if not ok(uid): await update.message.reply_text("⛔ Доступ заборонено."); return
    global ID_GROUP_CHAT
    d = load()
    if d.get("group_chat_id") and not ID_GROUP_CHAT: ID_GROUP_CHAT = d["group_chat_id"]
    await update.message.reply_text(f"☕ *КавоБот* запущений!\nВітаю, {uname(uid)}!",
                                    parse_mode="Markdown", reply_markup=main_kb())

async def cmd_balance(update, ctx):
    if not ok(update.effective_user.id): return
    await show_balance(update, load())

async def show_balance(update, d):
    s=d.get("stock",{}); sm=d.get("stats",{}).get("my",{}); sh=d.get("stats",{}).get("his",{})
    p=get_prices(d)
    stock_val=s.get("кава",0)*p["кава"]["buy"]+s.get("комплект",0)*p["комплект"]["buy"]
    profit=(sm.get("кава",0)*170+sm.get("комплект",0)*170+
            sh.get("кава",0)*70+sh.get("комплект",0)*70+
            (sh.get("молоко",0)+sh.get("айріш",0))*p["молоко"]["his"])
    await update.message.reply_text(
        f"💰 *БАЛАНС*\n{'─'*24}\n{bal_line(d['balance'])}\n\n"
        f"📦 *СКЛАД:*\n  ☕ Кава: {s.get('кава',0)} шт\n  📦 Комплекти: {s.get('комплект',0)} шт\n"
        f"  🥛 Молоко: {s.get('молоко',0)} шт\n  🍹 Айріш: {s.get('айріш',0)} шт\n"
        f"  💵 Вартість: {fm(stock_val)}\n\n"
        f"📊 *ПРОДАНО:*\n  Мої: кава {sm.get('кава',0)} | компл {sm.get('комплект',0)}\n"
        f"  Колі: кава {sh.get('кава',0)} | компл {sh.get('комплект',0)} | мол/айр {sh.get('молоко',0)+sh.get('айріш',0)}\n\n"
        f"💚 *Прибуток: {fm(profit)}*", parse_mode="Markdown")

async def cmd_report(update, ctx):
    if not ok(update.effective_user.id): return
    d=load(); txs=d.get("transactions",[])[-20:]
    if not txs: await update.message.reply_text("📊 Операцій немає."); return
    em={"sale":"☕","supply":"📥","payment":"💸","rent":"🏠","equipment":"🔧"}
    lines=["📊 *Останні 20 операцій*\n"]
    for t in reversed(txs):
        sign="+" if t["delta"]>0 else "-"
        lines.append(f"{em.get(t['type'],'•')} {t['date'][:10]} {sign}{fm(t['amount'])}  _{t['desc'][:35]}_")
    await update.message.reply_text("\n".join(lines), parse_mode="Markdown")

async def cmd_undo(update, ctx):
    if not ok(update.effective_user.id): return
    d=load(); txs=d.get("transactions",[])
    if not txs: await update.message.reply_text("❌ Немає операцій."); return
    last=txs.pop(); d["balance"]=round(d["balance"]-last["delta"],2)
    meta=last.get("meta",{})
    if last["type"]=="sale":
        for g,q in meta.get("items",{}).items():
            d["stock"][g]=d["stock"].get(g,0)+q
            sk="my" if meta.get("owner")=="volodymyr" else "his"
            d["stats"][sk][g]=max(0,d["stats"][sk].get(g,0)-q)
    elif last["type"]=="supply":
        for g,q in meta.get("items",{}).items():
            d["stock"][g]=max(0,d["stock"].get(g,0)-q)
    save(d)
    await update.message.reply_text(f"↩️ *Скасовано:* {last['desc']}\n{bal_line(d['balance'])}",
                                    parse_mode="Markdown")

async def cmd_export(update, ctx):
    if not ok(update.effective_user.id): return
    d=load(); txs=d.get("transactions",[])
    if not txs: await update.message.reply_text("❌ Немає даних."); return
    try:
        import openpyxl; os.makedirs(EXPORT_DIR,exist_ok=True)
        path=os.path.join(EXPORT_DIR,f"export_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx")
        wb=openpyxl.Workbook(); ws=wb.active; ws.title="Операції"
        ws.append(["ID","Дата","Тип","Опис","Сума","Зміна","Баланс"])
        for t in txs: ws.append([t["id"],t["date"][:16],t["type"],t["desc"],t["amount"],t["delta"],t["balance"]])
        wb.save(path)
        await update.message.reply_document(document=open(path,"rb"),filename=os.path.basename(path))
    except ImportError:
        lines=["Дата,Тип,Опис,Сума"]
        for t in txs: lines.append(f"{t['date'][:10]},{t['type']},{t['desc']},{t['amount']}")
        await update.message.reply_document(document="\n".join(lines).encode("utf-8-sig"),filename="export.csv")

async def cmd_stock(update, ctx):
    if not ok(update.effective_user.id): return
    d=load(); s=d.get("stock",{}); p=get_prices(d)
    val=s.get("кава",0)*p["кава"]["buy"]+s.get("комплект",0)*p["комплект"]["buy"]
    await update.message.reply_text(
        f"📦 *СКЛАД*\n  ☕ Кава: *{s.get('кава',0)}* шт\n  📦 Комплекти: *{s.get('комплект',0)}* шт\n"
        f"  🥛 Молоко: *{s.get('молоко',0)}* шт\n  🍹 Айріш: *{s.get('айріш',0)}* шт\n"
        f"  🥤 Ст.110: *{s.get('стакан110',0)}* шт\n  🥤 Ст.250: *{s.get('стакан250',0)}* шт\n\n"
        f"💵 Вартість: *{fm(val)}*", parse_mode="Markdown")

async def cmd_history(update, ctx):
    await cmd_report(update, ctx)

async def cmd_settings(update, ctx):
    if not ok(update.effective_user.id): return
    d=load(); p=get_prices(d)
    clients_debt=sum(c.get("debt",0) for c in d.get("clients",{}).values())
    await update.message.reply_text(
        f"⚙️ *НАЛАШТУВАННЯ*\n\n💲 *Ціни:*\n"
        f"  Кава: закупка {p['кава']['buy']} / продаж {p['кава']['sell']}\n"
        f"  Комплект: закупка {p['комплект']['buy']} / продаж {p['комплект']['sell']}\n"
        f"  Молоко комісія: {p['молоко']['his']} грн\n\n"
        f"👥 Борги клієнтів: {fm(clients_debt)}\n"
        f"📡 Група: {'✅ '+str(d.get('group_chat_id')) if d.get('group_chat_id') else '❌ не налаштована'}\n\n"
        "/setgroup — зробити цей чат груповим\n/setprice — змінити ціни\n"
        "/chart — графік\n/clients — борги\n/remind — нагадування\n"
        "/rent — оренда\n/export — Excel\n/undo — скасувати останнє",
        parse_mode="Markdown")

async def cmd_points(update, ctx):
    if not ok(update.effective_user.id): return
    d=load(); pts=d.get("points",{})
    if not pts:
        kb=InlineKeyboardMarkup([[InlineKeyboardButton("➕ Моя",callback_data="padd_volodymyr"),
                                   InlineKeyboardButton("➕ Колі",callback_data="padd_vygran")]])
        await update.message.reply_text("📍 Точок немає.",reply_markup=kb); return
    rows=[]
    for name,owner in pts.items():
        rows.append([InlineKeyboardButton(f"{'🏪' if owner=='volodymyr' else '🏬'} {name}",callback_data=f"pedit_{name}"),
                     InlineKeyboardButton("🔄",callback_data=f"pswap_{name}"),
                     InlineKeyboardButton("🗑️",callback_data=f"pdel_{name}")])
    rows.append([InlineKeyboardButton("➕ Моя",callback_data="padd_volodymyr"),
                 InlineKeyboardButton("➕ Колі",callback_data="padd_vygran")])
    my=[n for n,o in pts.items() if o=="volodymyr"]; his=[n for n,o in pts.items() if o=="vygran"]
    await update.message.reply_text(
        f"📍 *Точки*\nМої: {', '.join(my) or '—'}\nКолі: {', '.join(his) or '—'}",
        parse_mode="Markdown",reply_markup=InlineKeyboardMarkup(rows))

async def cmd_addpoint(update, ctx):
    uid=update.effective_user.id
    if not ok(uid): return
    args=ctx.args
    if len(args)<2: await update.message.reply_text("Використання: /addpoint <назва> <my|his>"); return
    name=" ".join(args[:-1]); owner="volodymyr" if args[-1].lower() in ("my","моя","мої") else "vygran"
    d=load(); d["points"][name]=owner; save(d)
    await update.message.reply_text(f"✅ Точку *{name}* додано.",parse_mode="Markdown")

async def ask_confirm(update, ctx, op_id, label, amount, meta):
    PENDING_CONFIRM[op_id]=meta
    kb=InlineKeyboardMarkup([[InlineKeyboardButton("✅ Підтвердити",callback_data=f"confirm_yes_{op_id}"),
                               InlineKeyboardButton("❌ Скасувати",callback_data=f"confirm_no_{op_id}")]])
    await update.message.reply_text(f"⚠️ *Велика сума*\n{label}: *{fm(amount)}*",
                                    parse_mode="Markdown",reply_markup=kb)

async def cmd_setprice(update, ctx):
    if not ok(update.effective_user.id): return
    d=load(); p=get_prices(d)
    kb=InlineKeyboardMarkup([[InlineKeyboardButton("☕ Кава",callback_data="sp_кава"),
                               InlineKeyboardButton("📦 Комплект",callback_data="sp_комплект")],
                              [InlineKeyboardButton("🥛 Молоко комісія",callback_data="sp_молоко")]])
    await update.message.reply_text(
        f"💲 *Ціни:*\n☕ Кава: {p['кава']['buy']}/{p['кава']['sell']}\n"
        f"📦 Комплект: {p['комплект']['buy']}/{p['комплект']['sell']}\n"
        f"🥛 Молоко/Айріш: {p['молоко']['his']} грн комісія\n\nОбери що змінити:",
        parse_mode="Markdown",reply_markup=kb)

async def cmd_receipt(update, ctx):
    if not ok(update.effective_user.id): return
    d=load(); receipt=d.get("_last_receipt")
    if not receipt: await update.message.reply_text("❌ Немає останнього продажу."); return
    kb=InlineKeyboardMarkup([[InlineKeyboardButton("📋 У борг клієнту",callback_data="receipt_debt"),
                               InlineKeyboardButton("✅ Закрити",callback_data="receipt_close")]])
    await update.message.reply_text(f"```\n{receipt}\n```\n_Скопіюй і надішли_",
                                    parse_mode="Markdown",reply_markup=kb)

async def cmd_clients(update, ctx):
    if not ok(update.effective_user.id): return
    d=load(); clients=d.get("clients",{})
    if not clients:
        kb=InlineKeyboardMarkup([[InlineKeyboardButton("➕ Додати",callback_data="cl_add")]])
        await update.message.reply_text("👥 Клієнтів немає.",reply_markup=kb); return
    total=sum(c.get("debt",0) for c in clients.values())
    debtors=[(n,c) for n,c in clients.items() if c.get("debt",0)>0]
    lines=["👥 *БОРГИ КЛІЄНТІВ*\n"]
    if debtors:
        lines.append("🔴 *З боргом:*")
        for n,c in sorted(debtors,key=lambda x:-x[1]["debt"]):
            try: days=(datetime.now()-datetime.fromisoformat(c.get("last_date",""))).days
            except: days=0
            lines.append(f"  • {n}: *{fm(c['debt'])}* ({days}д)")
    clean=[(n,c) for n,c in clients.items() if c.get("debt",0)<=0]
    if clean: lines.append("\n✅ "+", ".join(n for n,_ in clean))
    lines.append(f"\n💰 *Загалом: {fm(total)}*")
    rows=[[InlineKeyboardButton(f"{'👤' if c.get('debt',0)>0 else '✅'} {n} ({fm(c.get('debt',0))})",
                                callback_data=f"cl_view_{n}")] for n,c in clients.items()]
    rows.append([InlineKeyboardButton("➕ Новий клієнт",callback_data="cl_add")])
    await update.message.reply_text("\n".join(lines),parse_mode="Markdown",
                                    reply_markup=InlineKeyboardMarkup(rows))

async def cmd_remind_clients(update, ctx):
    if not ok(update.effective_user.id): return
    d=load(); debtors={n:c for n,c in d.get("clients",{}).items() if c.get("debt",0)>0}
    if not debtors: await update.message.reply_text("✅ Жодного боргу!"); return
    for name,c in sorted(debtors.items(),key=lambda x:-x[1]["debt"]):
        debt=c["debt"]
        try: days=(datetime.now()-datetime.fromisoformat(c.get("last_date",""))).days
        except: days=0
        reminder=f"Привіт, {name}!\n\nБорг за каву: {int(debt)} грн"
        if days: reminder+=f"\n(Брали {days} днів тому)"
        reminder+="\n\nРозрахуйся. Дякуємо! ☕"
        kb=InlineKeyboardMarkup([[InlineKeyboardButton("✅ Оплатив",callback_data=f"cl_pay_{name}")]])
        await update.message.reply_text(f"👤 *{name}* — {fm(debt)}\n```\n{reminder}\n```",
                                        parse_mode="Markdown",reply_markup=kb)

async def cmd_chart(update, ctx):
    if not ok(update.effective_user.id): return
    d=load(); txs=[t for t in d.get("transactions",[]) if t["type"]=="sale"]
    if not txs: await update.message.reply_text("📈 Ще немає продажів."); return
    daily=defaultdict(float)
    cutoff=(datetime.now()-timedelta(days=13)).strftime("%Y-%m-%d")
    for t in txs:
        if t["date"][:10]>=cutoff: daily[t["date"][:10]]+=t.get("amount",0)
    if not daily: await update.message.reply_text("Немає даних за 14 днів."); return
    days=sorted(daily.keys()); vals=[daily[dd] for dd in days]; max_v=max(vals) or 1
    rows=["*📈 Продажі за 14 днів*\n"]
    for day,val in zip(days,vals):
        n=int(val/max_v*8); rows.append(f"`{day[5:]}` {'█'*n}{'░'*(8-n)} {int(val/1000)}к")
    rows.append(f"\nМакс: {fm(max_v)}"); rows.append(f"Всього: {fm(sum(vals))}")
    await update.message.reply_text("\n".join(rows),parse_mode="Markdown")

async def cmd_rent(update, ctx):
    uid=update.effective_user.id
    if not ok(uid): return
    amount=RENT_AMOUNT
    if ctx.args:
        try: amount=int(ctx.args[0])
        except: pass
    if not amount:
        QUICK[uid]={"step":"exp_amount","etype":"rent","payer":"volodymyr"}
        await update.message.reply_text("🏠 Введи суму оренди:"); return
    QUICK[uid]={"step":"exp_payer","etype":"rent","amount":float(amount)}
    kb=InlineKeyboardMarkup([[InlineKeyboardButton("Я платив",callback_data="ep_volodymyr"),
                               InlineKeyboardButton("Коля платив",callback_data="ep_vygran")]])
    await update.message.reply_text(f"🏠 Оренда {fm(amount)} — хто платив?",reply_markup=kb)

def build_settlement_text(d):
    b=d.get("balance",0); s=d.get("stock",{}); sm=d.get("stats",{}).get("my",{}); sh=d.get("stats",{}).get("his",{})
    p=get_prices(d)
    profit=(sm.get("кава",0)*170+sm.get("комплект",0)*170+
            sh.get("кава",0)*70+sh.get("комплект",0)*70+
            (sh.get("молоко",0)+sh.get("айріш",0))*p["молоко"]["his"])
    stock_val=s.get("кава",0)*p["кава"]["buy"]+s.get("комплект",0)*p["комплект"]["buy"]
    em={"кава":"☕","комплект":"📦","молоко":"🥛","айріш":"🍹"}
    lines=["🤝 *РОЗРАХУНОК З КОЛЕЮ*","═"*24+"\n","💰 *БАЛАНС:*",
           "  ✅ Рахунки зведені" if abs(b)<1 else
           f"  🔴 Коля винен Тобі: *{fm(b)}*" if b>0 else
           f"  🔴 Ти винен Колі: *{fm(abs(b))}*",
           "","📦 *ЗАЛИШКИ:*"]
    has_s=False
    for good,qty in s.items():
        if qty>0: has_s=True; lines.append(f"  {em.get(good,'•')} {good}: *{qty} шт*")
    if not has_s: lines.append("  Порожній")
    if stock_val: lines.append(f"  💵 Вартість: *{fm(stock_val)}*")
    lines+=[f"","📊 *ПРОДАНО:*",
            f"  Мої: кава {sm.get('кава',0)} | компл {sm.get('комплект',0)}",
            f"  Колі: кава {sh.get('кава',0)} | компл {sh.get('комплект',0)}",
            "","─"*24,f"💚 *Прибуток: {fm(profit)}*","═"*24,
            "✅ *Нікому платити!*" if abs(b)<1 else
            f"👉 *Коля платить Тобі: {fm(b)}*" if b>0 else
            f"👉 *Ти платиш Колі: {fm(abs(b))}*"]
    return "\n".join(lines)

async def cmd_settlement(update, ctx):
    if not ok(update.effective_user.id): return
    d=load()
    kb=InlineKeyboardMarkup([[InlineKeyboardButton("📤 Надіслати Колі",callback_data="settle_send")],
                              [InlineKeyboardButton("✅ Оплачено — обнулити",callback_data="settle_paid")],
                              [InlineKeyboardButton("❌ Закрити",callback_data="settle_close")]])
    await update.message.reply_text(build_settlement_text(d),parse_mode="Markdown",reply_markup=kb)

async def cmd_setgroup(update, ctx):
    uid=update.effective_user.id
    if uid!=ID_VOLODYMYR: return
    global ID_GROUP_CHAT
    chat_id=update.effective_chat.id; d=load(); d["group_chat_id"]=chat_id; save(d)
    ID_GROUP_CHAT=chat_id
    await update.message.reply_text(f"✅ *Цей чат — груповий!*\nID: `{chat_id}`",
                                    parse_mode="Markdown")

async def on_photo(update, ctx):
    uid=update.effective_user.id
    if not ok(uid): return
    msg=await update.message.reply_text("🔍 Розпізнаю чек через AI...")
    photo=update.message.photo[-1]; f=await ctx.bot.get_file(photo.file_id)
    with tempfile.NamedTemporaryFile(suffix=".jpg",delete=False) as tmp:
        await f.download_to_drive(tmp.name)
        with open(tmp.name,"rb") as img_f: b64=base64.b64encode(img_f.read()).decode()
        os.unlink(tmp.name)
    info=await ai_photo(b64)
    if not info: await msg.edit_text("❌ Groq API не налаштований або помилка. Введи суму вручну."); return
    total=float(info.get("total",0)); date=info.get("date",""); items=info.get("items","")
    txt=f"🧾 *Чек розпізнано*\n\nСума: *{fm(total)}*"
    if date: txt+=f"\nДата: {date}"
    if items: txt+=f"\nТовари: {items}"
    kb=InlineKeyboardMarkup([[InlineKeyboardButton(f"💸 Витрата {fm(total)}",callback_data=f"photo_exp_{int(total)}")],
                              [InlineKeyboardButton("📥 Постачання",callback_data=f"photo_sup_{int(total)}")],
                              [InlineKeyboardButton("❌ Скасувати",callback_data="photo_cancel")]])
    await msg.edit_text(txt,parse_mode="Markdown",reply_markup=kb)

async def on_voice(update, ctx):
    uid=update.effective_user.id
    if not ok(uid): return
    if not GROQ_API_KEY: await update.message.reply_text("❌ Groq API не налаштований."); return
    msg=await update.message.reply_text("🎙 Розпізнаю голос...")
    f=await ctx.bot.get_file(update.message.voice.file_id)
    with tempfile.NamedTemporaryFile(suffix=".ogg",delete=False) as tmp:
        await f.download_to_drive(tmp.name); path=tmp.name
    text=await transcribe(path); os.unlink(path)
    if not text: await msg.edit_text("❌ Не вдалось розпізнати."); return
    await msg.edit_text(f"🎙 *Розпізнано:* {text}",parse_mode="Markdown")
    update.message.text=text
    await on_text(update,ctx)

GOOD_ALIASES={
    "кава":["кав","кава","кави","каву","кавою","coffee","кофе"],
    "комплект":["компл","комплект","комплекти","комплектів","комплекту","коплект","коплектів",
                "комплектов","комп","комлпект","комлп","кмплект","комплекта","комплектах"],
    "молоко":["молок","молоко","молока","молоку"],
    "айріш":["айріш","айриш","irish","айрiш"],
    "стакан110":["ст110","стакан110","110мл"],
    "стакан250":["ст250","стакан250","250мл"],
}
SUPPLY_KW=["привіз","поставив","постачання","прийшло","привезли","отримав","закупка",
            "прихід","приход","завіз","закупили","доставка","товар прийшов"]
PAY_KW=["передав","відав","віддав","оплатив","заплатив","розрахував","повернув","зп","зарплата"]
RENT_KW=["оренда","оренди","орендна плата"]
EQUIP_KW=["купюрник","принтер","обладнання","купив апарат"]

def normalize_good(word):
    w=word.lower().strip()
    for good,aliases in GOOD_ALIASES.items():
        if any(w.startswith(a) or a.startswith(w[:4]) for a in aliases if len(w)>=3):
            return good
    return None

def normalize_point(text, points):
    tl=text.lower()
    for point,owner in points.items():
        if point.lower() in tl: return point,owner
    return None,None

def parse_goods(tokens):
    items={}; i=0
    while i<len(tokens):
        tok=tokens[i]
        if re.match(r"[\+\-]?\d+",tok):
            num=abs(int(re.sub(r"[^\d]","",tok)))
            if i+1<len(tokens):
                g=normalize_good(tokens[i+1])
                if g: items[g]=items.get(g,0)+num; i+=2; continue
        else:
            g=normalize_good(tok)
            if g:
                num=None
                if i>0 and re.match(r"[\+\-]?\d+",tokens[i-1]):
                    num=abs(int(re.sub(r"[^\d]","",tokens[i-1])))
                elif i+1<len(tokens) and re.match(r"[\+\-]?\d+",tokens[i+1]):
                    num=abs(int(re.sub(r"[^\d]","",tokens[i+1]))); i+=1
                if num: items[g]=items.get(g,0)+num
        i+=1
    return items

async def smart_parse_free(text, d, update, ctx, uid):
    tl=text.lower().strip()
    tokens=re.findall(r"[\+\-]?\d+|[а-яіїєa-z]+\.?",tl)
    has_plus=bool(re.search(r"\+\d",tl))
    has_supply=any(kw in tl for kw in SUPPLY_KW)
    if has_plus or has_supply:
        items=parse_goods(tokens)
        m_grn=re.search(r"(\d+)\s*грн",tl)
        amount=float(m_grn.group(1)) if m_grn else 0.0
        if items:
            result=await do_supply(items,amount,d)
            await update.message.reply_text(result,parse_mode="Markdown"); return result
        if has_supply:
            QUICK[uid]={"step":"supply_text"}
            await update.message.reply_text("📥 Що привезли?\n_Напр: 20 комплектів 10 кава_",parse_mode="Markdown"); return ""
    m_amt=re.search(r"(\d+)",tl)
    if any(kw in tl for kw in PAY_KW) and m_amt:
        amount=float(m_amt.group()); payer="vygran" if any(w in tl for w in ["коля","він","партнер"]) else "volodymyr"
        if amount>=CONFIRM_SUM:
            await ask_confirm(update,ctx,f"pay_{uid}_{int(amount)}","Виплата",amount,{"op":"payment","payer":payer,"comment":text}); return ""
        result=await do_payment(amount,payer,d,text)
        await update.message.reply_text(result,parse_mode="Markdown"); return result
    if any(kw in tl for kw in RENT_KW) and m_amt:
        amount=float(m_amt.group()); payer="vygran" if any(w in tl for w in ["коля","він"]) else "volodymyr"
        result=await do_expense("rent",amount,payer,d,text)
        await update.message.reply_text(result,parse_mode="Markdown"); return result
    if any(kw in tl for kw in EQUIP_KW) and m_amt:
        amount=float(m_amt.group()); payer="vygran" if any(w in tl for w in ["коля","він"]) else "volodymyr"
        result=await do_expense("equipment",amount,payer,d,text)
        await update.message.reply_text(result,parse_mode="Markdown"); return result
    point,owner=normalize_point(tl,d.get("points",{}))
    if point:
        items=parse_goods(tokens)
        if items:
            result=await do_sale(point,owner,items,d)
            sale_kb=InlineKeyboardMarkup([[InlineKeyboardButton("🧾 Чек",callback_data="show_receipt"),
                                           InlineKeyboardButton("📋 У борг",callback_data="receipt_debt")]])
            await update.message.reply_text(result,parse_mode="Markdown",reply_markup=sale_kb); return result
        good_in_text=next((normalize_good(tok) for tok in tokens if normalize_good(tok)),None)
        if good_in_text:
            QUICK[uid]={"step":"qty","point":point,"owner":owner,"good":good_in_text}
            await update.message.reply_text(f"📍 *{point}* — скільки *{good_in_text}*?",parse_mode="Markdown"); return ""
        kb=InlineKeyboardMarkup([[InlineKeyboardButton("☕ Кава",callback_data=f"pt2_{point}_кава"),
                                   InlineKeyboardButton("📦 Комплект",callback_data=f"pt2_{point}_комплект")],
                                  [InlineKeyboardButton("🥛 Молоко",callback_data=f"pt2_{point}_молоко"),
                                   InlineKeyboardButton("🍹 Айріш",callback_data=f"pt2_{point}_айріш")]])
        await update.message.reply_text(f"📍 *{point}* — що продав?",parse_mode="Markdown",reply_markup=kb); return ""
    solo_good=solo_qty=None
    for i,tok in enumerate(tokens):
        g=normalize_good(tok)
        if g:
            solo_good=g
            if i>0 and re.match(r"\d+",tokens[i-1]): solo_qty=int(tokens[i-1])
            elif i+1<len(tokens) and re.match(r"\d+",tokens[i+1]): solo_qty=int(tokens[i+1])
            break
    if solo_good:
        QUICK[uid]={"step":"point","good":solo_good,"pending_qty":solo_qty}
        pts=list(d.get("points",{}).keys())
        if not pts: await update.message.reply_text("❌ Додай точки через 📍 Точки"); return ""
        kb=InlineKeyboardMarkup([[InlineKeyboardButton(p,callback_data=f"pt_{p}")] for p in pts])
        msg=f"📍 {solo_qty} {solo_good} — обери точку:" if solo_qty else f"📍 *{solo_good}* — обери точку:"
        await update.message.reply_text(msg,parse_mode="Markdown",reply_markup=kb); return ""
    return None

# ══════════════════════════════════════════════════════
# ШВИДКІ ПРОДАЖІ
# ══════════════════════════════════════════════════════
async def cmd_quick(update, ctx):
    """Меню швидких продажів — одним тапом"""
    uid = update.effective_user.id
    if not ok(uid): return
    d  = load()
    qs = d.get("quick_sales", [])
    if not qs:
        kb = InlineKeyboardMarkup([[InlineKeyboardButton("➕ Додати швидкий продаж", callback_data="qs_add")]])
        await update.message.reply_text("Швидкі продажі: немає. Додай через qs_add", reply_markup=kb)



    rows = []
    for i, qs_item in enumerate(qs):
        label = f"{qs_item['point']}: {' '.join(str(v)+' '+g for g,v in qs_item['items'].items())}"
        rows.append([InlineKeyboardButton(f"⚡ {label}", callback_data=f"qs_do_{i}")])
    rows.append([
        InlineKeyboardButton("➕ Додати", callback_data="qs_add"),
        InlineKeyboardButton("🗑️ Видалити", callback_data="qs_del"),
    ])
    await update.message.reply_text("Швидкі продажі — обери:", reply_markup=InlineKeyboardMarkup(rows))



# ══════════════════════════════════════════════════════
# ТОП ТОЧОК + ПОРІВНЯННЯ МІСЯЦІВ
# ══════════════════════════════════════════════════════
async def cmd_top(update, ctx):
    """Топ точок і порівняння місяців"""
    uid = update.effective_user.id
    if not ok(uid): return
    d   = load()
    txs = [t for t in d.get("transactions", []) if t["type"] == "sale"]
    if not txs:
        await update.message.reply_text("Ще немає продажів."); return

    now   = datetime.now()
    cur_m = now.strftime("%Y-%m")
    prev_m= (now.replace(day=1) - timedelta(days=1)).strftime("%Y-%m")

    # Топ точок (всі часи)
    point_sales = defaultdict(float)
    point_cur   = defaultdict(float)
    point_prev  = defaultdict(float)
    for t in txs:
        pt    = t.get("meta", {}).get("point", "?")
        amt   = t.get("amount", 0)
        month = t["date"][:7]
        point_sales[pt] += amt
        if month == cur_m:  point_cur[pt]  += amt
        if month == prev_m: point_prev[pt] += amt

    # Порівняння місяців
    cur_total  = sum(t["amount"] for t in txs if t["date"][:7] == cur_m)
    prev_total = sum(t["amount"] for t in txs if t["date"][:7] == prev_m)
    diff       = cur_total - prev_total
    arrow      = "📈" if diff >= 0 else "📉"

    nl   = chr(10)
    rows = [
        f"🏆 *Топ точок (всі часи)*",
        "",
    ]
    for i, (pt, amt) in enumerate(sorted(point_sales.items(), key=lambda x: -x[1])[:8], 1):
        cur  = point_cur.get(pt, 0)
        medal = ["🥇","🥈","🥉"][i-1] if i <= 3 else f"{i}."
        rows.append(f"  {medal} {pt}: *{fm(amt)}* (цей міс: {fm(cur)})")

    rows += [
        "",
        f"📅 *Порівняння місяців*",
        f"  Поточний ({cur_m}): *{fm(cur_total)}*",
        f"  Попередній ({prev_m}): *{fm(prev_total)}*",
        f"  {arrow} Різниця: *{fm(abs(diff))}* {'більше' if diff>=0 else 'менше'}",
    ]

    # Прогноз поточного місяця
    days_passed = now.day
    if days_passed > 0 and cur_total > 0:
        days_in_month = 30
        forecast = cur_total / days_passed * days_in_month
        rows += ["", f"🔮 *Прогноз на місяць: {fm(forecast)}*"]

    await update.message.reply_text(nl.join(rows), parse_mode="Markdown")

# ══════════════════════════════════════════════════════
# EXCEL ЗВІТ (місячний)
# ══════════════════════════════════════════════════════
async def cmd_monthly_excel(update, ctx):
    uid = update.effective_user.id
    if not ok(uid): return
    d   = load()
    txs = d.get("transactions", [])
    now = datetime.now()
    cur_m = now.strftime("%Y-%m")
    month_txs = [t for t in txs if t["date"][:7] == cur_m]
    if not month_txs:
        await update.message.reply_text(f"Немає даних за {cur_m}."); return

    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment
        os.makedirs(EXPORT_DIR, exist_ok=True)
        wb = openpyxl.Workbook()

        # Лист 1: Операції
        ws1 = wb.active; ws1.title = "Операції"
        hdr = Font(bold=True, color="FFFFFF")
        fill = PatternFill("solid", fgColor="2C3E50")
        headers = ["Дата", "Тип", "Опис", "Сума", "Баланс"]
        for col, h in enumerate(headers, 1):
            c = ws1.cell(1, col, h); c.font = hdr; c.fill = fill
            c.alignment = Alignment(horizontal="center")
        for row, t in enumerate(month_txs, 2):
            ws1.append([t["date"][:16], t["type"], t["desc"], t["amount"], t.get("balance", 0)])
        for col in ws1.columns:
            ws1.column_dimensions[col[0].column_letter].width = 20

        # Лист 2: Підсумок
        ws2 = wb.create_sheet("Підсумок")
        p   = get_prices(d)
        sm  = d["stats"]["my"]; sh = d["stats"]["his"]
        sales    = [t for t in month_txs if t["type"] == "sale"]
        payments = [t for t in month_txs if t["type"] == "payment"]
        profit   = (sm.get("кава",0)*170 + sm.get("комплект",0)*170 +
                    sh.get("кава",0)*70  + sh.get("комплект",0)*70  +
                    (sh.get("молоко",0)+sh.get("айріш",0))*100)

        summary = [
            [f"Звіт за {cur_m}", ""],
            ["Продажів", len(sales)],
            ["Сума продажів", sum(t["amount"] for t in sales)],
            ["Виплати", sum(t["amount"] for t in payments)],
            ["Баланс", d["balance"]],
            ["", ""],
            ["МОЇ ТОЧКИ", ""],
            ["Кава", sm.get("кава", 0)],
            ["Комплекти", sm.get("комплект", 0)],
            ["", ""],
            ["ТОЧКИ КОЛІ", ""],
            ["Кава", sh.get("кава", 0)],
            ["Комплекти", sh.get("комплект", 0)],
            ["", ""],
            ["ПРИБУТОК", profit],
        ]
        for row_data in summary:
            ws2.append(row_data)
        ws2.column_dimensions["A"].width = 25
        ws2.column_dimensions["B"].width = 15

        # Лист 3: Клієнти з боргами
        ws3 = wb.create_sheet("Борги клієнтів")
        ws3.append(["Клієнт", "Борг", "Останній раз"])
        for name, c in d.get("clients", {}).items():
            if c.get("debt", 0) > 0:
                last = c.get("last_date", "")[:10]
                ws3.append([name, c["debt"], last])

        fname = f"kavobiz_{cur_m}.xlsx"
        path  = os.path.join(EXPORT_DIR, fname)
        wb.save(path)
        await update.message.reply_document(
            document=open(path, "rb"), filename=fname,
            caption=f"📊 Місячний звіт {cur_m}")
    except ImportError:
        await update.message.reply_text("❌ openpyxl не встановлений. pip install openpyxl")
    except Exception as e:
        await update.message.reply_text(f"❌ Помилка: {e}")

# ══════════════════════════════════════════════════════
# ПІДТВЕРДЖЕННЯ ВІД КОЛІ
# ══════════════════════════════════════════════════════
async def cmd_kolja_confirm(update, ctx):
    """Коля може підтвердити або заперечити останню операцію"""
    uid = update.effective_user.id
    if uid != ID_VYGRAN: return
    d   = load()
    txs = d.get("transactions", [])
    if not txs:
        await update.message.reply_text("Операцій немає."); return
    last = txs[-1]
    if last.get("confirmed_by_kolja"):
        await update.message.reply_text(f"✅ Вже підтверджено: {last['desc']}"); return
    kb = InlineKeyboardMarkup([[
        InlineKeyboardButton("✅ Підтвердити", callback_data=f"kconf_yes_{last['id']}"),
        InlineKeyboardButton("❌ Заперечити",  callback_data=f"kconf_no_{last['id']}"),
    ]])
    await update.message.reply_text(
        "Остання операція від Володимира" + chr(10)
        + last["desc"] + chr(10)
        + fm(last["amount"]) + chr(10)
        + last["date"][:16] + chr(10) + "Підтверджуєш?",
        parse_mode="Markdown", reply_markup=kb)

async def on_text(update, ctx):
    uid=update.effective_user.id
    if not ok(uid): return
    text=update.message.text.strip(); d=load()
    state=QUICK.get(uid,{}); step=state.get("step")
    if text=="💰 Баланс":              await cmd_balance(update,ctx); return
    if text=="📊 Звіт":                await cmd_report(update,ctx); return
    if text=="📍 Точки":               await cmd_points(update,ctx); return
    if text=="👥 Борги клієнтів":      await cmd_clients(update,ctx); return
    if text=="💲 Змінити ціни":        await cmd_setprice(update,ctx); return
    if text=="🧾 Чек клієнту":         await cmd_receipt(update,ctx); return
    if text=="📈 Графік":              await cmd_chart(update,ctx); return
    if text=="🔔 Нагадати клієнтам":   await cmd_remind_clients(update,ctx); return
    if text=="🤝 Розрахунок з Колею":  await cmd_settlement(update,ctx); return
    if text=="⚙️ Налаштування":        await cmd_settings(update,ctx); return
    if text=="↩️ Скасувати":           await cmd_undo(update,ctx); return
    if text in ("☕ Продаж кави","📦 Продаж комплекту","🥛 Молоко / Айріш"):
        gmap={"☕ Продаж кави":"кава","📦 Продаж комплекту":"комплект","🥛 Молоко / Айріш":"молоко"}
        good=gmap[text]; pts=list(d.get("points",{}).keys())
        if not pts: await update.message.reply_text("❌ Додай точки через 📍 Точки"); return
        QUICK[uid]={"step":"point","good":good}
        kb=InlineKeyboardMarkup([[InlineKeyboardButton(p,callback_data=f"pt_{p}")] for p in pts])
        await update.message.reply_text(f"📍 Обери точку ({good}):",reply_markup=kb); return
    if text=="📥 Постачання":
        QUICK[uid]={"step":"supply_qty"}
        await update.message.reply_text(
            "📥 *Постачання — крок 1/3*\n\nСкільки привезли? Введи товари:\n"
            "_20 комплектів 10 кава_\nабо _+20 компл +10 кав 1500грн_",
            parse_mode="Markdown"); return
    if text=="💸 Виплата":
        kb=InlineKeyboardMarkup([[InlineKeyboardButton("💸 Ти → Колі",callback_data="pay_dir_volodymyr"),
                                   InlineKeyboardButton("💸 Коля → Тобі",callback_data="pay_dir_vygran")]])
        await update.message.reply_text("💸 Хто платить?",reply_markup=kb); return
    if text=="🏠 Оренда / Витрати":
        kb=InlineKeyboardMarkup([[InlineKeyboardButton("🏠 Оренда",callback_data="exp_rent"),
                                   InlineKeyboardButton("🔧 Обладнання",callback_data="exp_equipment")],
                                  [InlineKeyboardButton("🚚 Доставка",callback_data="exp_delivery")]])
        await update.message.reply_text("🏠 Тип витрати:",reply_markup=kb); return
    if step=="setprice_good":
        QUICK.pop(uid,None); good=state.get("good","кава"); d2=load()
        now=datetime.now().strftime("%d.%m.%Y %H:%M")
        if good=="молоко":
            m=re.search(r"\d+",text)
            if not m: await update.message.reply_text("❌ Введи число"); return
            commission=int(m.group())
            d2.setdefault("prices",{}).setdefault("молоко",{})["his"]=commission
            d2.setdefault("prices",{}).setdefault("айріш",{})["his"]=commission
            d2.setdefault("price_history",[]).append({"date":now,"good":"молоко/айріш","commission":commission})
            save(d2); await update.message.reply_text(f"✅ Комісія молоко/айріш: *{commission} грн*",parse_mode="Markdown"); return
        parts=re.findall(r"\d+",text)
        if len(parts)<2: await update.message.reply_text("❌ Введи два числа: _закупка продаж_",parse_mode="Markdown"); return
        buy,sell=int(parts[0]),int(parts[1])
        d2.setdefault("prices",{}).setdefault(good,{}).update({"buy":buy,"sell":sell,"my":buy,"his":sell-70})
        d2.setdefault("price_history",[]).append({"date":now,"good":good,"buy":buy,"sell":sell})
        save(d2); await update.message.reply_text(f"✅ *{good}*: {buy}/{sell} грн",parse_mode="Markdown"); return
    if step=="newclient_name":
        QUICK.pop(uid,None); name=text.strip().title(); d2=load()
        d2.setdefault("clients",{})[name]={"debt":0,"history":[]}; save(d2)
        kb=InlineKeyboardMarkup([[InlineKeyboardButton("➕ Борг",callback_data=f"cl_debt_{name}"),
                                   InlineKeyboardButton("👤 Переглянути",callback_data=f"cl_view_{name}")]])
        await update.message.reply_text(f"✅ Клієнт *{name}* доданий!",parse_mode="Markdown",reply_markup=kb); return
    if step=="debt_client_name":
        QUICK.pop(uid,None); name=text.strip().title(); sale=state.get("sale",{}); items=sale.get("items",{})
        prices=get_prices(d); total=sum(prices.get(g,{}).get("sell",0)*q for g,q in items.items())
        now=datetime.now(); d2=load()
        c=d2.setdefault("clients",{}).setdefault(name,{"debt":0,"history":[]})
        c["debt"]=c.get("debt",0)+total; c["last_date"]=now.isoformat()
        c.setdefault("history",[]).append({"type":"debt","amount":total,"date":now.strftime("%d.%m.%Y"),"note":"борг"})
        save(d2); await update.message.reply_text(f"✅ *{name}* +{fm(total)}\nЗагальний: {fm(c['debt'])}",parse_mode="Markdown"); return
    if step=="client_pay":
        QUICK.pop(uid,None); name=state.get("client",""); m=re.search(r"[\d.]+",text)
        if not m: await update.message.reply_text("❌ Введи суму"); return
        amount=float(m.group()); d2=load(); c=d2.get("clients",{}).get(name)
        if not c: await update.message.reply_text("❌ Клієнт не знайдений"); return
        c["debt"]=max(0,c.get("debt",0)-amount); c["last_date"]=datetime.now().isoformat()
        c.setdefault("history",[]).append({"type":"payment","amount":amount,"date":datetime.now().strftime("%d.%m.%Y"),"note":"оплата"})
        save(d2); await update.message.reply_text(f"✅ *{name}* заплатив {fm(amount)}\nЗалишок: *{fm(c['debt'])}*",parse_mode="Markdown"); return
    if step=="client_add_debt":
        QUICK.pop(uid,None); name=state.get("client",""); m=re.search(r"[\d.]+",text)
        if not m: await update.message.reply_text("❌ Введи суму"); return
        amount=float(m.group()); note=re.sub(r"[\d.]+","",text).strip() or "борг"; d2=load()
        c=d2.get("clients",{}).get(name)
        if not c: await update.message.reply_text("❌ Клієнт не знайдений"); return
        c["debt"]=c.get("debt",0)+amount; c["last_date"]=datetime.now().isoformat()
        c.setdefault("history",[]).append({"type":"debt","amount":amount,"date":datetime.now().strftime("%d.%m.%Y"),"note":note})
        save(d2); await update.message.reply_text(f"✅ *{name}* +{fm(amount)}\nЗагальний: *{fm(c['debt'])}*",parse_mode="Markdown"); return
    if step=="newpoint_name":
        QUICK.pop(uid,None); owner=state.get("owner","volodymyr"); name=text.strip()
        d.setdefault("points",{})[name]=owner; save(d)
        await update.message.reply_text(f"✅ Точку *{name}* ({'мою' if owner=='volodymyr' else 'Колі'}) додано!",parse_mode="Markdown"); return
    if step=="rename_point":
        QUICK.pop(uid,None); old_name=state.get("point",""); new_name=text.strip()
        if old_name in d.get("points",{}):
            owner=d["points"].pop(old_name); d["points"][new_name]=owner; save(d)
            await update.message.reply_text(f"✅ *{old_name}* → *{new_name}*",parse_mode="Markdown"); return
    if step=="pay_amount":
        QUICK.pop(uid,None); m=re.search(r"[\d.]+",text)
        if not m: await update.message.reply_text("❌ Введи суму"); return
        amount=float(m.group()); payer=state["payer"]
        if amount>=CONFIRM_SUM:
            await ask_confirm(update,ctx,f"pay_{uid}_{int(amount)}","Виплата",amount,
                              {"op":"payment","payer":payer,"comment":"виплата"}); return
        result=await do_payment(amount,payer,d)
        await update.message.reply_text(result,parse_mode="Markdown"); await check_debt(ctx,d); return
    if step=="exp_amount":
        QUICK.pop(uid,None); m=re.search(r"[\d.]+",text)
        if not m: await update.message.reply_text("❌ Введи суму"); return
        result=await do_expense(state.get("etype","rent"),float(m.group()),state.get("payer","volodymyr"),d)
        await update.message.reply_text(result,parse_mode="Markdown"); await check_debt(ctx,d); return
    if step=="qty":
        QUICK.pop(uid,None); m=re.search(r"\d+",text)
        if not m: await update.message.reply_text("❌ Введи кількість"); return
        qty=int(m.group()); result=await do_sale(state["point"],state["owner"],{state["good"]:qty},d)
        sale_kb=InlineKeyboardMarkup([[InlineKeyboardButton("🧾 Чек",callback_data="show_receipt"),
                                        InlineKeyboardButton("📋 У борг",callback_data="receipt_debt")]])
        await update.message.reply_text(result,parse_mode="Markdown",reply_markup=sale_kb)
        await check_debt(ctx,d); return
    if step in ("supply_text", "supply_qty"):
        tokens2=re.findall(r"[\+\-]?\d+|[а-яіїєa-z]+\.?",text.lower())
        items=parse_goods(tokens2); m_grn=re.search(r"(\d+)\s*грн",text.lower())
        amount=float(m_grn.group(1)) if m_grn else 0.0
        if not items:
            await update.message.reply_text("❌ Не зрозумів. Спробуй: _20 комплектів 10 кава_",parse_mode="Markdown")
            return
        # Ask prices for each good
        goods_list = [g for g in items if g in ("кава","комплект")]
        if goods_list:
            QUICK[uid]={"step":"supply_price","items":items,"amount":amount,"price_idx":0}
            good = goods_list[0]
            p = get_prices(d)
            cur_buy = p.get(good,{}).get("buy",0)
            cur_sell= p.get(good,{}).get("sell",0)
            await update.message.reply_text(
                f"📥 *Постачання — крок 2/3*\n\n"
                f"*{good.capitalize()}* ({items[good]} шт)\n"
                f"Поточні ціни: закупка *{cur_buy}* / продаж *{cur_sell}*\n\n"
                f"Введи нові ціни через пробіл: _закупка продаж_\n"
                f"Або натисни /skip щоб залишити поточні",
                parse_mode="Markdown")
        else:
            QUICK.pop(uid,None)
            result=await do_supply(items,amount,d)
            await update.message.reply_text(result,parse_mode="Markdown")
        return

    if step=="supply_price":
        items     = state.get("items",{})
        amount    = state.get("amount",0.0)
        goods_list= [g for g in items if g in ("кава","комплект")]
        idx       = state.get("price_idx",0)
        good      = goods_list[idx] if idx < len(goods_list) else None
        # Parse prices (or /skip)
        if good and text.strip() not in ("/skip","skip","пропустити","-"):
            nums = re.findall(r"\d+", text)
            if len(nums) >= 2:
                buy, sell = int(nums[0]), int(nums[1])
                # Save new prices
                d2 = load()
                d2.setdefault("prices",{}).setdefault(good,{}).update({
                    "buy": buy, "sell": sell,
                    "my":  buy,
                    "his": sell - d2.get("prices",{}).get(good,{}).get("commission",70),
                })
                d2.setdefault("price_history",[]).append({
                    "date": datetime.now().strftime("%d.%m.%Y %H:%M"),
                    "good": good, "buy": buy, "sell": sell,
                    "note": "при постачанні"
                })
                save(d2)
                QUICK[uid]["prices_changed"] = QUICK[uid].get("prices_changed",[]) + [f"{good}: {buy}/{sell}"]
        # Move to next good
        next_idx = idx + 1
        if next_idx < len(goods_list):
            QUICK[uid]["price_idx"] = next_idx
            next_good = goods_list[next_idx]
            d2 = load()
            p  = get_prices(d2)
            cur_buy  = p.get(next_good,{}).get("buy",0)
            cur_sell = p.get(next_good,{}).get("sell",0)
            await update.message.reply_text(
                f"📥 *Постачання — {next_good.capitalize()}* ({items[next_good]} шт)\n"
                f"Поточні: закупка *{cur_buy}* / продаж *{cur_sell}*\n\n"
                f"Нові ціни (_закупка продаж_) або /skip:",
                parse_mode="Markdown")
        else:
            # All prices done — record supply
            QUICK.pop(uid,None)
            d2     = load()
            result = await do_supply(items, amount, d2)
            changed= QUICK.get(uid,{}).get("prices_changed",[])
            if changed:
                result += "\n\n💲 *Ціни оновлено:* " + ", ".join(changed)
            await update.message.reply_text(result, parse_mode="Markdown")
        return
    # /skip during supply_price — treat as skip
    if text.strip().startswith("/skip") and step=="supply_price":
        # Re-run same handler with "skip" text
        state2 = QUICK.get(uid,{}).copy()
        QUICK[uid] = {**state2}
        # Simulate skip by calling the step again  
        # Handled above — just fall through won't work, so inline:
        items     = state.get("items",{})
        amount    = state.get("amount",0.0)
        goods_list= [g for g in items if g in ("кава","комплект")]
        idx       = state.get("price_idx",0)
        next_idx  = idx + 1
        if next_idx < len(goods_list):
            QUICK[uid]["price_idx"] = next_idx
            next_good = goods_list[next_idx]
            p = get_prices(d)
            await update.message.reply_text(
                f"{next_good.capitalize()} ({items[next_good]} шт)\n"
                f"Ціни: {p[next_good]['buy']}/{p[next_good]['sell']}\n/skip або _закупка продаж_:",
                parse_mode="Markdown")
        else:
            QUICK.pop(uid,None)
            result_s = await do_supply(items, amount, d)
            await update.message.reply_text(result_s, parse_mode="Markdown")
        return

    result=await smart_parse_free(text,d,update,ctx,uid)
    if result is not None: return
    if GROQ_API_KEY:
        msg=await update.message.reply_text("🤔 Думаю...")
        parsed=await ai_parse(text); ptype=parsed.get("type","unknown")
        point=parsed.get("point",""); items=parsed.get("items",{}); amount=float(parsed.get("amount",0))
        payer=parsed.get("payer","volodymyr"); comment=parsed.get("comment","")
        result_text=None
        if ptype=="sale" and point and items:
            owner=d.get("points",{}).get(point,"volodymyr"); result_text=await do_sale(point,owner,items,d)
        elif ptype=="supply" and items: result_text=await do_supply(items,amount,d,comment)
        elif ptype=="payment" and amount:
            if amount>=CONFIRM_SUM:
                await msg.delete()
                await ask_confirm(update,ctx,f"pay_{uid}_{int(amount)}","Виплата",amount,{"op":"payment","payer":payer,"comment":comment}); return
            result_text=await do_payment(amount,payer,d,comment)
        elif ptype in ("rent","equipment","delivery") and amount:
            result_text=await do_expense(ptype,amount,payer,d,comment)
        if result_text: await msg.edit_text(result_text,parse_mode="Markdown"); await check_debt(ctx,d)
        else: await msg.edit_text("❓ Не зрозумів — скористайся кнопками 👇",reply_markup=main_kb())
    else:
        await update.message.reply_text("❓ Не зрозумів — скористайся кнопками 👇",reply_markup=main_kb())

async def on_callback(update, ctx):
    q=update.callback_query; await q.answer()
    uid=q.from_user.id; cb=q.data; d=load()
    if cb.startswith("confirm_yes_"):
        op_id=cb[12:]; meta=PENDING_CONFIRM.pop(op_id,{})
        if not meta: await q.edit_message_text("❌ Операція застаріла."); return
        if meta.get("op")=="payment":
            result=await do_payment(float(op_id.split("_")[-1]),meta["payer"],d,meta.get("comment",""))
            await q.edit_message_text(result,parse_mode="Markdown")
        await check_debt(ctx,d); return
    if cb.startswith("confirm_no_"):
        PENDING_CONFIRM.pop(cb[11:],None); await q.edit_message_text("❌ Скасовано."); return
    if cb.startswith("photo_exp_"):
        amount=float(cb[10:]); QUICK[uid]={"step":"exp_payer","etype":"rent","amount":amount}
        kb=InlineKeyboardMarkup([[InlineKeyboardButton("Я платив",callback_data="ep_volodymyr"),
                                   InlineKeyboardButton("Коля платив",callback_data="ep_vygran")]])
        await q.edit_message_text(f"💸 Витрата {fm(amount)} — хто платив?",reply_markup=kb); return
    if cb.startswith("photo_sup_"):
        result=await do_supply({},float(cb[10:]),d,"фото чеку")
        await q.edit_message_text(result,parse_mode="Markdown"); return
    if cb=="photo_cancel": await q.delete_message(); return
    if cb.startswith("sp_"):
        good=cb[3:]; QUICK[uid]={"step":"setprice_good","good":good}; p=get_prices(d).get(good,{})
        if good=="молоко": await q.edit_message_text(f"Комісія молоко зараз: *{p.get('his',100)} грн*\nВведи нову:",parse_mode="Markdown")
        else: await q.edit_message_text(f"*{good}*: зараз {p.get('buy')}/{p.get('sell')}\nВведи: _закупка продаж_",parse_mode="Markdown")
        return
    if cb=="show_receipt":
        receipt=d.get("_last_receipt","")
        if receipt:
            await q.edit_message_text(f"```\n{receipt}\n```\n_Скопіюй_",parse_mode="Markdown",
                reply_markup=InlineKeyboardMarkup([[InlineKeyboardButton("📋 У борг",callback_data="receipt_debt"),
                                                    InlineKeyboardButton("✅ Закрити",callback_data="receipt_close")]])); return
    if cb=="receipt_debt":
        sale=d.get("_last_sale",{}); QUICK[uid]={"step":"debt_client_name","sale":sale}
        existing="\n".join(f"• {n}" for n in d.get("clients",{}).keys())
        await q.edit_message_text("👤 Введи ім'я клієнта:"+(f"\n\nІснуючі:\n{existing}" if existing else "")); return
    if cb=="receipt_close": await q.delete_message(); return
    if cb=="settle_send":
        text=build_settlement_text(load())
        try:
            await ctx.bot.send_message(ID_VYGRAN,text+"\n\n_Надіслано Тобі_",parse_mode="Markdown")
            await q.edit_message_text(text+"\n\n✅ *Надіслано Колі!*",parse_mode="Markdown")
        except Exception as e: await q.edit_message_text(f"❌ Помилка: {e}")
        return
    if cb=="settle_paid":
        d2=load(); old_bal=d2.get("balance",0)
        if abs(old_bal)>=1:
            delta=-old_bal if old_bal>0 else abs(old_bal)
            add_tx(d2,"payment",f"Розрахунок: {fm(abs(old_bal))}",abs(old_bal),delta,{"settlement":True})
        d2["stats"]={"my":{"кава":0,"комплект":0,"молоко":0,"айріш":0},"his":{"кава":0,"комплект":0,"молоко":0,"айріш":0}}
        save(d2)
        stock_txt="\n".join(f"  {g}: {q2} шт" for g,q2 in d2["stock"].items() if q2>0) or "  Порожній"
        await q.edit_message_text(f"✅ *Розрахунок завершено!*\nБаланс обнулено.\n\n📦 Склад:\n{stock_txt}",parse_mode="Markdown")
        try: await ctx.bot.send_message(ID_VYGRAN,"✅ *Розрахунок підтверджено*\nБаланс обнулено.",parse_mode="Markdown")
        except: pass
        return
    if cb=="settle_close": await q.edit_message_text("❌ Скасовано."); return
    if cb.startswith("pay_dir_"):
        payer=cb[8:]; QUICK[uid]={"step":"pay_amount","payer":payer}
        label="Ти → Колі" if payer=="volodymyr" else "Коля → Тобі"
        await q.edit_message_text(f"💸 {label}\nВведи суму:"); return
    if cb.startswith("ep_"):
        payer=cb[3:]; state=QUICK.get(uid,{}); etype=state.get("etype","rent"); amount=state.get("amount",0)
        QUICK.pop(uid,None); result=await do_expense(etype,float(amount),payer,d)
        await q.edit_message_text(result,parse_mode="Markdown"); await check_debt(ctx,d); return
    if cb.startswith("exp_"):
        etype=cb[4:]; QUICK[uid]={"step":"exp_payer","etype":etype}
        kb=InlineKeyboardMarkup([[InlineKeyboardButton("Я платив",callback_data="ep_volodymyr"),
                                   InlineKeyboardButton("Коля платив",callback_data="ep_vygran")]])
        await q.edit_message_text(f"🏠 {etype} — хто платив?",reply_markup=kb); return
    if cb=="show_balance": await show_balance(update,load()); return
    if cb.startswith("pt_"):
        point=cb[3:]; owner=d.get("points",{}).get(point,"volodymyr"); state=QUICK.get(uid,{})
        good=state.get("good"); pending_qty=state.get("pending_qty")
        if not good:
            QUICK.pop(uid,None)
            kb=InlineKeyboardMarkup([[InlineKeyboardButton("☕ Кава",callback_data=f"pt2_{point}_кава"),
                                       InlineKeyboardButton("📦 Комплект",callback_data=f"pt2_{point}_комплект")],
                                      [InlineKeyboardButton("🥛 Молоко",callback_data=f"pt2_{point}_молоко"),
                                       InlineKeyboardButton("🍹 Айріш",callback_data=f"pt2_{point}_айріш")]])
            await q.edit_message_text(f"📍 *{point}* — що продав?",parse_mode="Markdown",reply_markup=kb); return
        if pending_qty:
            QUICK.pop(uid,None); result=await do_sale(point,owner,{good:pending_qty},d)
            sale_kb=InlineKeyboardMarkup([[InlineKeyboardButton("🧾 Чек",callback_data="show_receipt"),
                                           InlineKeyboardButton("📋 У борг",callback_data="receipt_debt")]])
            await q.edit_message_text(result,parse_mode="Markdown",reply_markup=sale_kb)
            await check_debt(ctx,d); return
        QUICK[uid]={"step":"qty","point":point,"owner":owner,"good":good}
        await q.edit_message_text(f"📍 *{point}* — скільки *{good}*?",parse_mode="Markdown"); return
    if cb.startswith("pt2_"):
        parts=cb[4:].split("_",1); point,good=parts[0],parts[1] if len(parts)>1 else "кава"
        owner=d.get("points",{}).get(point,"volodymyr")
        QUICK[uid]={"step":"qty","point":point,"owner":owner,"good":good}
        await q.edit_message_text(f"📍 *{point}* — скільки *{good}*?",parse_mode="Markdown"); return
    if cb=="cl_add": QUICK[uid]={"step":"newclient_name"}; await q.edit_message_text("👤 Введи ім'я клієнта:"); return
    if cb.startswith("cl_view_"):
        name=cb[8:]; c=d.get("clients",{}).get(name,{}); debt=c.get("debt",0); hist=c.get("history",[])[-5:]
        lines=[f"👤 *{name}*",f"Борг: *{fm(debt)}*",""]
        if hist:
            lines.append("📋 *Останні операції:*")
            for h in reversed(hist): lines.append(f"  {h['date'][:10]}  {'+' if h['type']=='debt' else '-'}{fm(h['amount'])}  {h.get('note','')}")
        kb=InlineKeyboardMarkup([[InlineKeyboardButton("💸 Погасити",callback_data=f"cl_pay_{name}"),
                                   InlineKeyboardButton("➕ Борг",callback_data=f"cl_debt_{name}")],
                                  [InlineKeyboardButton("🗑️ Видалити",callback_data=f"cl_del_{name}"),
                                   InlineKeyboardButton("◀️ Назад",callback_data="cl_back")]])
        await q.edit_message_text("\n".join(lines),parse_mode="Markdown",reply_markup=kb); return
    if cb=="cl_back":
        d2=load(); total=sum(c.get("debt",0) for c in d2.get("clients",{}).values())
        rows=[[InlineKeyboardButton(f"{'👤' if c.get('debt',0)>0 else '✅'} {n} ({fm(c.get('debt',0))})",callback_data=f"cl_view_{n}")] for n,c in d2.get("clients",{}).items()]
        rows.append([InlineKeyboardButton("➕ Новий",callback_data="cl_add")])
        await q.edit_message_text(f"👥 *Клієнти*\nЗагалом: {fm(total)}",parse_mode="Markdown",reply_markup=InlineKeyboardMarkup(rows)); return
    if cb.startswith("cl_pay_"):
        name=cb[7:]; QUICK[uid]={"step":"client_pay","client":name}
        debt=d.get("clients",{}).get(name,{}).get("debt",0)
        await q.edit_message_text(f"💸 *{name}*\nБорг: *{fm(debt)}*\nВведи суму оплати:",parse_mode="Markdown"); return
    if cb.startswith("cl_debt_"):
        name=cb[8:]; QUICK[uid]={"step":"client_add_debt","client":name}
        await q.edit_message_text(f"➕ *{name}* — введи суму (і опис):\n_Напр: 350 2 кави_",parse_mode="Markdown"); return
    if cb.startswith("cl_del_"):
        name=cb[7:]; d.get("clients",{}).pop(name,None); save(d)
        await q.edit_message_text(f"🗑️ *{name}* видалено.",parse_mode="Markdown"); return
    if cb.startswith("padd_"):
        owner=cb[5:]; QUICK[uid]={"step":"newpoint_name","owner":owner}
        await q.edit_message_text(f"📍 Введи назву нової точки ({'моя' if owner=='volodymyr' else 'Колі'}):"); return
    if cb.startswith("pedit_"):
        point=cb[6:]; QUICK[uid]={"step":"rename_point","point":point}
        await q.edit_message_text(f"✏️ Нова назва для *{point}*:",parse_mode="Markdown"); return
    if cb.startswith("pswap_"):
        point=cb[6:]; old=d.get("points",{}).get(point,"volodymyr")
        new="vygran" if old=="volodymyr" else "volodymyr"
        d["points"][point]=new; save(d)
        await q.edit_message_text(f"🔄 *{point}* тепер {'моя' if new=='volodymyr' else 'Колі'}",parse_mode="Markdown"); return
    if cb.startswith("pdel_"):
        point=cb[5:]
        kb=InlineKeyboardMarkup([[InlineKeyboardButton("✅ Видалити",callback_data=f"pdelok_{point}"),
                                   InlineKeyboardButton("❌ Ні",callback_data="pdelno")]])
        await q.edit_message_text(f"🗑️ Видалити *{point}*?",parse_mode="Markdown",reply_markup=kb); return
    if cb.startswith("pdelok_"):
        point=cb[7:]; d.get("points",{}).pop(point,None); save(d)
        await q.edit_message_text(f"🗑️ *{point}* видалено.",parse_mode="Markdown"); return
    if cb=="pdelno": await q.edit_message_text("❌ Скасовано."); return

async def check_debt(ctx, d):
    b=d.get("balance",0)
    if abs(b)>=DEBT_ALERT:
        await notify_all(ctx.bot,f"⚠️ *Великий борг!*\n{bal_line(b)}\nРозрахуйся.")

async def daily_jobs(ctx):
    now=datetime.now(); d=load()
    if now.weekday()==0:
        today=now.strftime("%Y-%m-%d")
        if d.get("weekly_sent")!=today:
            txs=d.get("transactions",[]); week=(now-timedelta(days=7)).strftime("%Y-%m-%d")
            sales=[t for t in txs if t["type"]=="sale" and t["date"][:10]>=week]
            await notify_all(ctx.bot,f"📊 *Тижневий звіт*\n\nПродажів: {len(sales)}\nСума: {fm(sum(t['amount'] for t in sales))}\n\n{bal_line(d['balance'])}")
            d["weekly_sent"]=today; save(d)
    if RENT_AMOUNT>0 and now.day==RENT_DAY:
        month=now.strftime("%Y-%m")
        if d.get("rent_reminded")!=month:
            await notify_all(ctx.bot,f"🏠 *Нагадування про оренду*\nСьогодні {now.day}-е.\nОренда: *{fm(RENT_AMOUNT)}*\n/rent — записати.")
            d["rent_reminded"]=month; save(d)
    if now.weekday()==4:
        debtors=[(n,c["debt"]) for n,c in d.get("clients",{}).items() if c.get("debt",0)>0]
        if debtors:
            total=sum(v for _,v in debtors)
            lines=["👥 *Борги клієнтів (п'ятниця)*\n"]
            for n,debt in sorted(debtors,key=lambda x:-x[1]): lines.append(f"  • {n}: {fm(debt)}")
            lines.append(f"\n💰 Всього: *{fm(total)}*\n/clients — переглянути")
            await notify_all(ctx.bot,"\n".join(lines))

from aiohttp import web as aioWeb
WEB_PORT=int(os.environ.get("PORT",8080))

async def web_index(request):
    p=os.path.join(os.path.dirname(__file__),"dashboard.html")
    if not os.path.exists(p): return aioWeb.Response(text="dashboard.html not found",status=404)
    return aioWeb.Response(text=open(p,encoding="utf-8").read(),content_type="text/html",charset="utf-8")

async def api_get_data(request):
    if request.headers.get("X-Token","")!=BOT_TOKEN[:20]: return aioWeb.json_response({"error":"unauthorized"},status=401)
    return aioWeb.json_response(load())

async def api_post_data(request):
    if request.headers.get("X-Token","")!=BOT_TOKEN[:20]: return aioWeb.json_response({"error":"unauthorized"},status=401)
    try:
        data=await request.json(); save(data); return aioWeb.json_response({"ok":True})
    except Exception as e: return aioWeb.json_response({"error":str(e)},status=400)

async def api_health(request):
    return aioWeb.json_response({"ok":True,"time":datetime.now().isoformat()})

def make_web_app():
    app=aioWeb.Application()
    app.router.add_get("/",web_index); app.router.add_get("/api/data",api_get_data)
    app.router.add_post("/api/data",api_post_data); app.router.add_get("/health",api_health)
    return app

async def run_bot(tg_app):
    await tg_app.initialize(); await tg_app.start()
    await tg_app.updater.start_polling(drop_pending_updates=True)

async def run_web():
    app=make_web_app(); runner=aioWeb.AppRunner(app); await runner.setup()
    site=aioWeb.TCPSite(runner,"0.0.0.0",WEB_PORT); await site.start()
    logger.info(f"Веб: http://0.0.0.0:{WEB_PORT}")

async def main_async():
    os.makedirs(EXPORT_DIR,exist_ok=True)
    global ID_GROUP_CHAT
    try:
        d0=load()
        if d0.get("group_chat_id") and not ID_GROUP_CHAT: ID_GROUP_CHAT=d0["group_chat_id"]
    except Exception: pass
    tg_app=Application.builder().token(BOT_TOKEN).build()
    for cmd,fn in [("start",cmd_start),("balance",cmd_balance),("report",cmd_report),
                   ("export",cmd_export),("undo",cmd_undo),("points",cmd_points),
                   ("addpoint",cmd_addpoint),("stock",cmd_stock),("history",cmd_history),
                   ("settings",cmd_settings),("setprice",cmd_setprice),("receipt",cmd_receipt),
                   ("clients",cmd_clients),("remind",cmd_remind_clients),("chart",cmd_chart),
                   ("rent",cmd_rent),("settlement",cmd_settlement),("setgroup",cmd_setgroup)]:
        tg_app.add_handler(CommandHandler(cmd,fn))
    tg_app.add_handler(MessageHandler(filters.TEXT & ~filters.COMMAND, on_text))
    tg_app.add_handler(MessageHandler(filters.VOICE & ~filters.COMMAND, on_voice))
    tg_app.add_handler(MessageHandler(filters.PHOTO & ~filters.COMMAND, on_photo))
    tg_app.add_handler(CallbackQueryHandler(on_callback))
    tg_app.job_queue.run_daily(daily_jobs,time=datetime.strptime("09:00","%H:%M").time())
    await run_web(); await run_bot(tg_app)
    try: await asyncio.Event().wait()
    except (KeyboardInterrupt,SystemExit):
        await tg_app.updater.stop(); await tg_app.stop(); await tg_app.shutdown()

def main():
    asyncio.run(main_async())

if __name__ == "__main__":
    main()
